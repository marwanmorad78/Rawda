from decimal import Decimal, InvalidOperation
from datetime import datetime, time
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.views import LoginView, LogoutView
from django.core.files.base import ContentFile
from django.core.files.utils import validate_file_name
from django.db import transaction
from django.db.models import ProtectedError, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, TemplateView, UpdateView
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, UnidentifiedImageError

from apps.catalog.models import (
    Category,
    Company,
    Product,
    ProductCompany,
    ProductCompanyOption,
    ProductOption,
)
from apps.core.localization import DEFAULT_LANGUAGE, format_syp, localize_instance, localize_queryset
from apps.core.models import (
    CenterStatus,
    CustomerOrder,
    CustomerOrderItem,
    CustomerProfile,
    DeliveryArea,
    DeliverySubArea,
    SiteSettings,
)
from apps.core.pricing import (
    get_active_site_settings,
    set_product_display_price,
    set_promotion_display_price,
)
from apps.dashboard.forms import (
    CategoryForm,
    CenterStatusForm,
    CustomerAccessPasswordForm,
    CustomerPasswordChangeForm,
    DeliveryAreaExcelUploadForm,
    DeliveryAreaForm,
    DeliverySubAreaFormSet,
    DollarPriceForm,
    ManagerLoginForm,
    OrderAcceptForm,
    ProductExcelUploadForm,
    ProductCompanyFormSet,
    ProductCompanyOptionFormSet,
    ProductOptionFormSet,
    ProductForm,
    PromotionForm,
    SiteSettingsForm,
)
from apps.dashboard.localization import get_dashboard_strings
from apps.dashboard.mixins import DashboardLocalizationMixin, StaffRequiredMixin
from apps.core.order_status import attach_order_display, attach_orders_display, suggested_expected_minutes
from apps.core.services import mark_order_accepted, sync_center_status_and_auto_accept_waiting_orders
from apps.promotions.models import Promotion


PRODUCT_EXCEL_HEADERS = [
    "product_name",
    "category",
    "short_description",
    "description",
    "price",
    "product_type",
    "price_link_mode",
    "is_price_linked_to_dollar",
    "sold_by_weight_mode",
    "sold_by_weight",
    "has_options",
    "unit_label",
    "is_available",
    "is_featured",
    "image",
    "external_image_url",
    "sku",
]
PRODUCT_EXCEL_REQUIRED_HEADERS = ["product_name", "category", "price"]
PRODUCT_OPTION_EXCEL_HEADERS = [
    "product_sku",
    "option_name",
    "price",
    "is_default",
    "is_available",
    "display_order",
]
PRODUCT_OPTION_EXCEL_REQUIRED_HEADERS = ["product_sku", "option_name", "price"]
PRODUCT_COMPANY_EXCEL_HEADERS = [
    "company_id",
    "company_name",
    "logo",
    "external_logo_url",
    "display_order",
    "is_active",
]
PRODUCT_COMPANY_EXCEL_REQUIRED_HEADERS = ["company_id", "company_name"]
PRODUCT_COMPANY_OPTION_EXCEL_HEADERS = [
    "product_sku",
    "company_id",
    "option_name",
    "price",
    "is_available",
    "display_order",
]
PRODUCT_COMPANY_OPTION_EXCEL_REQUIRED_HEADERS = ["product_sku", "company_id", "option_name", "price"]
CATEGORY_IMAGE_EXCEL_HEADERS = [
    "category",
    "type",
    "parent",
    "name",
    "slug",
    "image",
    "external_image_url",
]
MAX_EXCEL_IMAGE_BYTES = 10 * 1024 * 1024
BOOLEAN_COLUMNS = ["is_price_linked_to_dollar", "sold_by_weight", "has_options", "is_available", "is_featured"]
PRODUCT_MODE_COLUMNS = {
    "price_link_mode": {Product.BEHAVIOR_INHERIT, Product.BEHAVIOR_CUSTOM},
    "sold_by_weight_mode": {Product.BEHAVIOR_INHERIT, Product.BEHAVIOR_CUSTOM},
}
DELIVERY_AREA_EXCEL_HEADERS = ["area_name", "has_sub_areas", "delivery_fee", "display_order", "is_active"]
DELIVERY_AREA_EXCEL_REQUIRED_HEADERS = ["area_name"]
DELIVERY_SUB_AREA_EXCEL_HEADERS = ["area_name", "sub_area_name", "delivery_fee", "display_order", "is_active"]
DELIVERY_SUB_AREA_EXCEL_REQUIRED_HEADERS = ["area_name", "sub_area_name", "delivery_fee"]
SALES_REPORT_STATUSES = [
    CustomerOrder.STATUS_ACCEPTED,
    CustomerOrder.STATUS_BEING_PREPARED,
    CustomerOrder.STATUS_OUT_FOR_DELIVERY,
    CustomerOrder.STATUS_READY_TO_PICKUP,
    CustomerOrder.STATUS_DONE,
]


def get_dashboard_site_settings():
    site_settings, _ = SiteSettings.objects.get_or_create(
        pk=1,
        defaults={
            "store_name": "Al Rawda Center",
            "store_name_ar": "مركز الروضة",
            "about_text": "Update your store information here.",
            "about_text_ar": "حدث معلومات المتجر من هنا.",
            "address": "Damascus",
            "address_ar": "دمشق",
            "primary_phone": "+963-000-000-000",
            "hero_title": "Fresh products for every home",
            "hero_title_ar": "منتجات طازجة لكل بيت",
        },
    )
    return site_settings


def normalize_excel_header(value):
    return str(value or "").strip().lower()


def normalize_excel_text(value):
    if value is None:
        return ""
    return str(value).strip()


def get_excel_image_search_paths(excel_file):
    paths = []
    file_name = getattr(excel_file, "name", "")
    if file_name:
        workbook_path = Path(file_name).expanduser()
        if workbook_path.is_absolute():
            paths.append(workbook_path.parent)
    temporary_file_path = getattr(excel_file, "temporary_file_path", None)
    if callable(temporary_file_path):
        try:
            paths.append(Path(temporary_file_path()).resolve().parent)
        except (OSError, ValueError):
            pass
    paths.extend(
        [
            Path(settings.PROJECT_ROOT),
            Path(settings.MEDIA_ROOT),
            Path.cwd(),
        ]
    )
    return list(dict.fromkeys(path.resolve() for path in paths))


def validate_excel_image_content(content, filename):
    if not content:
        raise ValueError("The image file is empty.")
    if len(content) > MAX_EXCEL_IMAGE_BYTES:
        raise ValueError("The image file is larger than 10 MB.")
    try:
        with Image.open(BytesIO(content)) as image:
            image.verify()
            image_format = (image.format or "").lower()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ValueError("The file is not a valid image.") from exc

    safe_name = validate_file_name(
        Path(filename).name or f"excel-image.{image_format}",
        allow_relative_path=False,
    )
    if not Path(safe_name).suffix and image_format:
        safe_name = f"{safe_name}.{image_format}"
    return {"name": safe_name, "content": content}


def load_excel_image(value, search_paths):
    source = normalize_excel_text(value)
    if not source:
        return None

    parsed = urlparse(source)
    if parsed.scheme.lower() in {"http", "https"}:
        request = Request(source, headers={"User-Agent": "AlRawdaExcelImporter/1.0"})
        try:
            with urlopen(request, timeout=15) as response:
                content_length = response.headers.get("Content-Length")
                if content_length and int(content_length) > MAX_EXCEL_IMAGE_BYTES:
                    raise ValueError("The image file is larger than 10 MB.")
                content = response.read(MAX_EXCEL_IMAGE_BYTES + 1)
                final_url = response.geturl()
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError("The image URL could not be downloaded.") from exc
        filename = Path(unquote(urlparse(final_url).path)).name or "excel-image"
        return validate_excel_image_content(content, filename)

    source_path = Path(source).expanduser()
    candidates = [source_path] if source_path.is_absolute() else [
        base_path / source_path
        for base_path in search_paths
    ]
    image_path = next((candidate for candidate in candidates if candidate.is_file()), None)
    if image_path is None:
        raise ValueError(f"Image file not found: {source}")
    try:
        content = image_path.read_bytes()
    except OSError as exc:
        raise ValueError(f"Image file could not be read: {source}") from exc
    return validate_excel_image_content(content, image_path.name)


def save_imported_image(instance, field_name, image_data):
    if not image_data:
        return
    image_field = getattr(instance, field_name)
    old_name = image_field.name
    image_field.save(
        image_data["name"],
        ContentFile(image_data["content"]),
        save=False,
    )
    instance.__class__.objects.filter(pk=instance.pk).update(
        **{
            field_name: image_field.name,
            "updated_at": timezone.now(),
        }
    )
    if old_name and old_name != image_field.name:
        transaction.on_commit(
            lambda storage=image_field.storage, name=old_name: storage.delete(name)
        )


def merge_legacy_company_records(canonical_company):
    legacy_companies = list(
        Company.objects.filter(
            code__startswith="legacy-",
            name__iexact=canonical_company.name,
        ).exclude(pk=canonical_company.pk)
    )
    for legacy_company in legacy_companies:
        if not canonical_company.logo and legacy_company.logo:
            canonical_company.logo = legacy_company.logo
        if not canonical_company.external_logo_url and legacy_company.external_logo_url:
            canonical_company.external_logo_url = legacy_company.external_logo_url
        canonical_company.save()

        for source_relation in list(legacy_company.products.all()):
            target_relation = ProductCompany.objects.filter(
                product=source_relation.product,
                company=canonical_company,
            ).first()
            if target_relation is None:
                source_relation.company = canonical_company
                source_relation.save(update_fields=["company", "updated_at"])
                continue

            for source_option in list(source_relation.options.all()):
                target_option = target_relation.options.filter(
                    name__iexact=source_option.name
                ).first()
                if target_option is None:
                    source_option.company = target_relation
                    source_option.save(update_fields=["company", "updated_at"])
                else:
                    CustomerOrderItem.objects.filter(selected_option=source_option).update(
                        selected_option=target_option
                    )
                    source_option.delete()
            CustomerOrderItem.objects.filter(company=source_relation).update(
                company=target_relation
            )
            source_relation.delete()

        legacy_company.delete()


def parse_excel_boolean(value, default=False):
    if value is None or value == "":
        return default, None
    if isinstance(value, bool):
        return value, None
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1", "available"}:
        return True, None
    if normalized in {"false", "no", "n", "0", "unavailable"}:
        return False, None
    return default, "Use true or false."


def parse_excel_decimal(value, allow_decimals=False):
    if value is None or value == "":
        return None, "This field is required."
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None, "Enter a valid number."
    if parsed < 0:
        return None, "Price cannot be negative."
    if not allow_decimals and parsed != parsed.to_integral_value():
        return None, "Enter a whole number without decimals."
    return parsed, None


def parse_excel_integer(value, default=0):
    if value is None or value == "":
        return default, None
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        return default, "Enter a whole number."
    if parsed < 0:
        return default, "Enter zero or a positive number."
    return parsed, None


def parse_excel_product_type(value, default=Product.PRODUCT_TYPE_NORMAL):
    normalized = normalize_excel_text(value).lower()
    if not normalized:
        return default, None
    aliases = {
        "normal": Product.PRODUCT_TYPE_NORMAL,
        "normal product": Product.PRODUCT_TYPE_NORMAL,
        "company_grouped": Product.PRODUCT_TYPE_COMPANY_GROUPED,
        "company grouped": Product.PRODUCT_TYPE_COMPANY_GROUPED,
        "company/brand grouped product": Product.PRODUCT_TYPE_COMPANY_GROUPED,
        "brand": Product.PRODUCT_TYPE_COMPANY_GROUPED,
        "company": Product.PRODUCT_TYPE_COMPANY_GROUPED,
    }
    product_type = aliases.get(normalized)
    if product_type is None:
        return default, "Use normal or company_grouped."
    return product_type, None


def build_sheet_rows(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize_excel_header(value) for value in next(rows, [])]
    header_positions = {header: index for index, header in enumerate(headers) if header}
    data_rows = []
    for row_number, row_values in enumerate(rows, start=2):
        row = {
            header: row_values[index] if index < len(row_values) else None
            for header, index in header_positions.items()
        }
        if any(value not in (None, "") for value in row.values()):
            data_rows.append((row_number, row))
    return header_positions, data_rows


def get_product_excel_category_lookup():
    lookup = {}
    duplicate_names = set()
    for category in Category.objects.select_related("parent").all():
        slug_key = (category.slug or "").strip().lower()
        name_key = (category.name or "").strip().lower()
        name_ar_key = (category.name_ar or "").strip().lower()
        if slug_key:
            lookup[slug_key] = category
            if category.parent_id and category.parent.slug:
                lookup[f"{category.parent.slug.strip().lower()}/{slug_key}"] = category
        if name_key:
            if name_key in lookup and lookup[name_key].pk != category.pk:
                duplicate_names.add(name_key)
            else:
                lookup[name_key] = category
        if name_ar_key:
            if name_ar_key in lookup and lookup[name_ar_key].pk != category.pk:
                duplicate_names.add(name_ar_key)
            else:
                lookup[name_ar_key] = category
        if category.parent_id:
            parent_names = [category.parent.name, category.parent.name_ar, category.parent.slug]
            child_names = [category.name, category.name_ar, category.slug]
            for parent_name in parent_names:
                for child_name in child_names:
                    parent_key = (parent_name or "").strip().lower()
                    child_key = (child_name or "").strip().lower()
                    if parent_key and child_key:
                        lookup[f"{parent_key}/{child_key}"] = category
                        lookup[f"{parent_key} / {child_key}"] = category
                        lookup[f"{parent_key} > {child_key}"] = category
    return lookup, duplicate_names


def format_category_path(category, language=DEFAULT_LANGUAGE):
    name = category.name_ar if language == "ar" and category.name_ar else category.name
    if category.parent_id:
        parent_name = category.parent.name_ar if language == "ar" and category.parent.name_ar else category.parent.name
        return f"{parent_name} / {name}"
    return name


def get_product_manage_context(language, product_form=None, upload_form=None, excel_errors=None):
    return {
        "dashboard_ui": get_dashboard_strings(language),
        "form": product_form or ProductForm(language=language),
        "upload_form": upload_form or ProductExcelUploadForm(language=language),
        "excel_errors": excel_errors or [],
    }


def get_product_list_context(language, upload_form=None, excel_errors=None, search_query="", filters=None):
    filters = filters or {}
    site_settings = get_active_site_settings() or get_dashboard_site_settings()
    items_queryset = Product.objects.select_related("category").all().order_by("name")
    if search_query:
        items_queryset = items_queryset.filter(
            Q(name__icontains=search_query)
            | Q(name_ar__icontains=search_query)
            | Q(short_description__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(category__name_ar__icontains=search_query)
            | Q(sku__icontains=search_query)
        )
    name_filter = (filters.get("name") or "").strip()
    category_filter = (filters.get("category") or "").strip()
    pricing_type_filter = (filters.get("pricing_type") or "").strip()
    status_filter = (filters.get("status") or "").strip()
    if name_filter:
        items_queryset = items_queryset.filter(
            Q(name__icontains=name_filter)
            | Q(name_ar__icontains=name_filter)
        )
    if category_filter.isdigit():
        items_queryset = items_queryset.filter(category_id=int(category_filter))
    if status_filter == "available":
        items_queryset = items_queryset.filter(is_available=True)
    elif status_filter == "unavailable":
        items_queryset = items_queryset.filter(is_available=False)
    items = list(items_queryset)
    for item in items:
        localize_instance(item.category, language, ["name", "description"])
        localize_instance(item, language, ["name", "short_description", "description", "unit_label"])
        set_product_display_price(item, language, site_settings)
    if pricing_type_filter:
        items = [
            item
            for item in items
            if (
                (pricing_type_filter == "sold_by_weight" and item.effective_sold_by_weight)
                or (pricing_type_filter == "dollar_priced" and item.effective_is_price_linked_to_dollar)
                or (
                    pricing_type_filter == "regular"
                    and not item.effective_sold_by_weight
                    and not item.effective_is_price_linked_to_dollar
                )
            )
        ]
    categories = list(Category.objects.select_related("parent").all().order_by("parent__name", "name"))
    for category in categories:
        localize_instance(category, language, ["name", "description"])
    return {
        "dashboard_ui": get_dashboard_strings(language),
        "items": items,
        "upload_form": upload_form or ProductExcelUploadForm(language=language),
        "excel_errors": excel_errors or [],
        "search_query": search_query,
        "product_filters": {
            "name": name_filter,
            "category": category_filter,
            "pricing_type": pricing_type_filter,
            "status": status_filter,
        },
        "product_filter_categories": categories,
    }


def get_delivery_area_manage_context(language, area_form=None, upload_form=None, excel_errors=None):
    return {
        "dashboard_ui": get_dashboard_strings(language),
        "form": area_form or DeliveryAreaForm(language=language),
        "upload_form": upload_form or DeliveryAreaExcelUploadForm(language=language),
        "excel_errors": excel_errors or [],
        "items": DeliveryArea.objects.prefetch_related("sub_areas").all(),
    }


def validate_product_excel_workbook(excel_file, language=DEFAULT_LANGUAGE):
    dashboard_ui = get_dashboard_strings(language)
    # Validate the entire workbook before importing so partial product batches are avoided.
    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
    except Exception:
        return [], [dashboard_ui["excel_read_error"]]

    products_sheet = workbook["Products"] if "Products" in workbook.sheetnames else workbook.active
    header_positions, product_sheet_rows = build_sheet_rows(products_sheet)
    option_header_positions = {}
    option_sheet_rows = []
    option_product_skus = set()
    company_header_positions = {}
    company_sheet_rows = []
    company_option_header_positions = {}
    company_option_sheet_rows = []
    company_option_product_skus = set()
    category_sheet_rows = []
    image_search_paths = get_excel_image_search_paths(excel_file)
    if "Options" in workbook.sheetnames:
        option_header_positions, option_sheet_rows = build_sheet_rows(workbook["Options"])
        if "product_sku" in option_header_positions:
            option_product_skus = {
                normalize_excel_text(row.get("product_sku")).lower()
                for _, row in option_sheet_rows
                if normalize_excel_text(row.get("product_sku"))
            }
    if "Companies" in workbook.sheetnames:
        company_header_positions, company_sheet_rows = build_sheet_rows(workbook["Companies"])
    if "CompanyOptions" in workbook.sheetnames:
        company_option_header_positions, company_option_sheet_rows = build_sheet_rows(workbook["CompanyOptions"])
        if "product_sku" in company_option_header_positions:
            company_option_product_skus = {
                normalize_excel_text(row.get("product_sku")).lower()
                for _, row in company_option_sheet_rows
                if normalize_excel_text(row.get("product_sku"))
            }
    if "Categories" in workbook.sheetnames:
        _, category_sheet_rows = build_sheet_rows(workbook["Categories"])
    missing_headers = [
        header for header in PRODUCT_EXCEL_REQUIRED_HEADERS if header not in header_positions
    ]
    if missing_headers:
        return [], [f"Missing required column: {', '.join(missing_headers)}"]

    category_lookup, duplicate_names = get_product_excel_category_lookup()
    valid_categories = []
    valid_rows = []
    valid_options = []
    valid_companies = []
    valid_company_options = []
    errors = []
    seen_import_keys = set()

    for row_number, row in category_sheet_rows:
        image_source = normalize_excel_text(row.get("image") or row.get("cover_image"))
        external_image_url = normalize_excel_text(row.get("external_image_url"))
        if not image_source and not external_image_url:
            continue
        category_key = normalize_excel_text(row.get("slug") or row.get("category")).lower()
        category = category_lookup.get(category_key)
        if category is None:
            errors.append(
                f"Categories row {row_number}: category '{category_key}' does not exist."
            )
            continue
        image_data = None
        if image_source:
            try:
                image_data = load_excel_image(image_source, image_search_paths)
            except ValueError as exc:
                errors.append(f"Categories row {row_number}: image: {exc}")
                continue
        valid_categories.append(
            {
                "category": category,
                "image": image_data,
                "external_image_url": external_image_url,
            }
        )

    for row_number, row in product_sheet_rows:
        row_errors = []
        product_name = normalize_excel_text(row.get("product_name"))
        category_value = normalize_excel_text(row.get("category"))
        sku = normalize_excel_text(row.get("sku"))

        if not product_name:
            row_errors.append("product_name is required.")

        category = None
        category_key = category_value.lower()
        if not category_value:
            row_errors.append("category is required.")
        elif category_key in duplicate_names:
            row_errors.append("category name is duplicated; use the category slug instead.")
        else:
            category = category_lookup.get(category_key)
            if not category:
                row_errors.append(f"category '{category_value}' does not exist.")

        parsed_booleans = {}
        for column in BOOLEAN_COLUMNS:
            default = column == "is_available"
            parsed_booleans[column], boolean_error = parse_excel_boolean(row.get(column), default)
            if boolean_error:
                row_errors.append(f"{column}: {boolean_error}")

        product_type_default = (
            Product.PRODUCT_TYPE_COMPANY_GROUPED
            if sku and sku.lower() in company_option_product_skus
            else Product.PRODUCT_TYPE_NORMAL
        )
        product_type, product_type_error = parse_excel_product_type(
            row.get("product_type"),
            product_type_default,
        )
        if product_type_error:
            row_errors.append(f"product_type: {product_type_error}")

        has_options = (
            product_type != Product.PRODUCT_TYPE_COMPANY_GROUPED
            and (parsed_booleans["has_options"] or bool(sku and sku.lower() in option_product_skus))
        )
        price, price_error = parse_excel_decimal(row.get("price"), allow_decimals=True)
        if price_error and (has_options or product_type == Product.PRODUCT_TYPE_COMPANY_GROUPED) and row.get("price") in (None, ""):
            price = Decimal("0")
        elif price_error:
            row_errors.append(f"price: {price_error}")

        parsed_modes = {}
        for column, choices in PRODUCT_MODE_COLUMNS.items():
            value = normalize_excel_text(row.get(column)) or Product.BEHAVIOR_INHERIT
            if value not in choices:
                row_errors.append(f"{column}: use inherit or custom.")
            parsed_modes[column] = value

        import_key = ("sku", sku.lower()) if sku else ("name", getattr(category, "id", None), product_name.lower())
        if import_key in seen_import_keys:
            row_errors.append("This product appears more than once in this file.")
        seen_import_keys.add(import_key)

        if sku and category is not None and product_name:
            product_by_sku = Product.objects.filter(sku__iexact=sku).first()
            product_by_name = Product.objects.filter(category=category, name__iexact=product_name).first()
            if product_by_sku and product_by_name and product_by_sku.pk != product_by_name.pk:
                row_errors.append(
                    f"sku '{sku}' belongs to a different product than '{product_name}'."
                )

        image_data = None
        image_source = normalize_excel_text(row.get("image") or row.get("primary_image"))
        if image_source and not row_errors:
            try:
                image_data = load_excel_image(image_source, image_search_paths)
            except ValueError as exc:
                row_errors.append(f"image: {exc}")

        if row_errors:
            errors.append(f"Row {row_number}: {' '.join(row_errors)}")
            continue

        row_data = {
            "category": category,
            "name": product_name,
            "short_description": normalize_excel_text(row.get("short_description")),
            "description": normalize_excel_text(row.get("description")),
            "price": price,
            "product_type": product_type,
            "price_link_mode": parsed_modes["price_link_mode"],
            "is_price_linked_to_dollar": parsed_booleans["is_price_linked_to_dollar"],
            "sold_by_weight_mode": parsed_modes["sold_by_weight_mode"],
            "sold_by_weight": parsed_booleans["sold_by_weight"],
            "has_options": has_options,
            "unit_label": normalize_excel_text(row.get("unit_label")) or "per item",
            "is_available": parsed_booleans["is_available"],
            "is_featured": parsed_booleans["is_featured"],
            "external_image_url": normalize_excel_text(row.get("external_image_url")),
            "sku": sku,
            "_image": image_data,
        }
        row_data["_import_keys"] = [import_key]
        if sku:
            row_data["_import_keys"].append(("name", getattr(category, "id", None), product_name.lower()))
        valid_rows.append(row_data)

    known_product_keys = {
        import_key
        for valid_row in valid_rows
        for import_key in valid_row.get("_import_keys", [])
    }
    imported_product_types_by_sku = {
        row["sku"].strip().lower(): row["product_type"]
        for row in valid_rows
        if row.get("sku")
    }

    if "Options" in workbook.sheetnames:
        if option_sheet_rows:
            missing_option_headers = [
                header for header in PRODUCT_OPTION_EXCEL_REQUIRED_HEADERS if header not in option_header_positions
            ]
            if missing_option_headers:
                errors.append(f"Options sheet missing required column: {', '.join(missing_option_headers)}")

        seen_option_keys = set()
        for row_number, row in option_sheet_rows:
            row_errors = []
            product_sku = normalize_excel_text(row.get("product_sku"))
            option_name = normalize_excel_text(row.get("option_name"))

            if not product_sku:
                row_errors.append("product_sku is required.")
            if not option_name:
                row_errors.append("option_name is required.")

            price, price_error = parse_excel_decimal(row.get("price"), allow_decimals=True)
            if price_error:
                row_errors.append(f"price: {price_error}")

            is_default, boolean_error = parse_excel_boolean(row.get("is_default"), False)
            if boolean_error:
                row_errors.append(f"is_default: {boolean_error}")
            is_available, available_error = parse_excel_boolean(row.get("is_available"), True)
            if available_error:
                row_errors.append(f"is_available: {available_error}")

            display_order, display_order_error = parse_excel_integer(row.get("display_order"), 0)
            if display_order_error:
                row_errors.append(f"display_order: {display_order_error}")

            product_key = ("sku", product_sku.lower()) if product_sku else None
            if product_key and product_key not in known_product_keys and not Product.objects.filter(sku__iexact=product_sku).exists():
                row_errors.append(f"product_sku '{product_sku}' does not match an imported or existing product.")
            if product_sku and imported_product_types_by_sku.get(product_sku.lower()) == Product.PRODUCT_TYPE_COMPANY_GROUPED:
                row_errors.append("Use CompanyOptions for company_grouped products, not Options.")

            option_key = (product_key, option_name.lower())
            if product_key and option_name:
                if option_key in seen_option_keys:
                    row_errors.append("This option appears more than once for the same product in this file.")
                seen_option_keys.add(option_key)

            if row_errors:
                errors.append(f"Options row {row_number}: {' '.join(row_errors)}")
                continue

            valid_options.append(
                {
                    "product_key": product_key,
                    "name": option_name,
                    "price": price,
                    "is_default": is_default,
                    "is_available": is_available,
                    "display_order": display_order,
                }
            )

    known_company_ids = {
        code.lower()
        for code in Company.objects.values_list("code", flat=True)
    }
    if "Companies" in workbook.sheetnames:
        if company_sheet_rows:
            missing_company_headers = [
                header for header in PRODUCT_COMPANY_EXCEL_REQUIRED_HEADERS if header not in company_header_positions
            ]
            if missing_company_headers:
                errors.append(f"Companies sheet missing required column: {', '.join(missing_company_headers)}")

        seen_company_ids = set()
        for row_number, row in company_sheet_rows:
            row_errors = []
            company_id = normalize_excel_text(row.get("company_id"))
            company_name = normalize_excel_text(row.get("company_name"))

            if not company_id:
                row_errors.append("company_id is required.")
            if not company_name:
                row_errors.append("company_name is required.")

            display_order, display_order_error = parse_excel_integer(row.get("display_order"), 0)
            if display_order_error:
                row_errors.append(f"display_order: {display_order_error}")
            is_active, active_error = parse_excel_boolean(row.get("is_active"), True)
            if active_error:
                row_errors.append(f"is_active: {active_error}")

            company_key = company_id.lower()
            if company_id:
                if company_key in seen_company_ids:
                    row_errors.append("This company_id appears more than once in this file.")
                seen_company_ids.add(company_key)
                known_company_ids.add(company_key)

            logo_data = None
            logo_source = normalize_excel_text(row.get("logo") or row.get("logo_url"))
            if logo_source and not row_errors:
                try:
                    logo_data = load_excel_image(logo_source, image_search_paths)
                except ValueError as exc:
                    row_errors.append(f"logo: {exc}")

            if row_errors:
                errors.append(f"Companies row {row_number}: {' '.join(row_errors)}")
                continue

            valid_companies.append(
                {
                    "company_id": company_id,
                    "name": company_name,
                    "logo": logo_data,
                    "external_logo_url": normalize_excel_text(row.get("external_logo_url")),
                    "order": display_order,
                    "is_active": is_active,
                }
            )

    if "CompanyOptions" in workbook.sheetnames:
        if company_option_sheet_rows:
            missing_company_option_headers = [
                header for header in PRODUCT_COMPANY_OPTION_EXCEL_REQUIRED_HEADERS
                if header not in company_option_header_positions
            ]
            if missing_company_option_headers:
                errors.append(
                    f"CompanyOptions sheet missing required column: {', '.join(missing_company_option_headers)}"
                )

        seen_company_option_keys = set()
        for row_number, row in company_option_sheet_rows:
            row_errors = []
            product_sku = normalize_excel_text(row.get("product_sku"))
            company_id = normalize_excel_text(row.get("company_id"))
            option_name = normalize_excel_text(row.get("option_name"))
            product_key = ("sku", product_sku.lower()) if product_sku else None
            company_key = (product_key, company_id.lower()) if product_key and company_id else None

            if not product_sku:
                row_errors.append("product_sku is required.")
            elif product_key not in known_product_keys and not Product.objects.filter(sku__iexact=product_sku).exists():
                row_errors.append(f"product_sku '{product_sku}' does not match an imported or existing product.")
            if not company_id:
                row_errors.append("company_id is required.")
            elif company_id.lower() not in known_company_ids:
                row_errors.append(f"company_id '{company_id}' does not match a Companies row.")
            if not option_name:
                row_errors.append("option_name is required.")

            price, price_error = parse_excel_decimal(row.get("price"), allow_decimals=True)
            if price_error:
                row_errors.append(f"price: {price_error}")

            is_available, available_error = parse_excel_boolean(row.get("is_available"), True)
            if available_error:
                row_errors.append(f"is_available: {available_error}")

            display_order, display_order_error = parse_excel_integer(row.get("display_order"), 0)
            if display_order_error:
                row_errors.append(f"display_order: {display_order_error}")

            option_key = (company_key, option_name.lower())
            if company_key and option_name:
                if option_key in seen_company_option_keys:
                    row_errors.append("This company option appears more than once for the same company in this file.")
                seen_company_option_keys.add(option_key)

            if row_errors:
                errors.append(f"CompanyOptions row {row_number}: {' '.join(row_errors)}")
                continue

            valid_company_options.append(
                {
                    "product_key": product_key,
                    "company_id": company_id,
                    "name": option_name,
                    "price": price,
                    "is_available": is_available,
                    "order": display_order,
                }
            )

    if not valid_rows and not errors:
        errors.append("The uploaded workbook has no product rows.")

    return {
        "categories": valid_categories,
        "products": valid_rows,
        "options": valid_options,
        "companies": valid_companies,
        "company_options": valid_company_options,
    }, errors


def find_existing_product_for_import(row):
    sku = (row.get("sku") or "").strip()
    if sku:
        existing_product = Product.objects.filter(sku__iexact=sku).first()
        if existing_product is not None:
            return existing_product
    return Product.objects.filter(category=row["category"], name__iexact=row["name"]).first()


def merge_duplicate_products_for_import():
    duplicates_removed = 0
    seen_keys = set()

    for product in Product.objects.select_related("category").order_by("id"):
        keys = []
        if product.sku:
            keys.append(("sku", product.sku.strip().lower()))
        keys.append(("name", product.category_id, product.name.strip().lower()))

        if any(key in seen_keys for key in keys):
            product.delete()
            duplicates_removed += 1
            continue

        seen_keys.update(keys)

    return duplicates_removed


def resolve_imported_product(product_lookup, product_key):
    product = product_lookup.get(product_key)
    if product is not None:
        return product
    if product_key and product_key[0] == "sku":
        return Product.objects.filter(sku__iexact=product_key[1]).first()
    if product_key and product_key[0] == "name":
        return Product.objects.filter(category_id=product_key[1], name__iexact=product_key[2]).first()
    return None


def import_product_excel_rows(valid_rows):
    # Existing products are matched by SKU when present, otherwise by category and product name.
    product_rows = valid_rows["products"] if isinstance(valid_rows, dict) else valid_rows
    category_rows = valid_rows.get("categories", []) if isinstance(valid_rows, dict) else []
    option_rows = valid_rows.get("options", []) if isinstance(valid_rows, dict) else []
    company_rows = valid_rows.get("companies", []) if isinstance(valid_rows, dict) else []
    company_option_rows = valid_rows.get("company_options", []) if isinstance(valid_rows, dict) else []
    created = 0
    updated = 0
    now = timezone.now()

    with transaction.atomic():
        for row in category_rows:
            category = row["category"]
            Category.objects.filter(pk=category.pk).update(
                external_image_url=row["external_image_url"],
                updated_at=now,
            )
            save_imported_image(category, "cover_image", row.get("image"))

        merge_duplicate_products_for_import()
        product_lookup = {}
        for row in product_rows:
            product_data = row.copy()
            import_keys = product_data.pop("_import_keys", [])
            image_data = product_data.pop("_image", None)
            existing_product = find_existing_product_for_import(product_data)
            if existing_product:
                update_data = product_data.copy()
                if not update_data["sku"]:
                    update_data.pop("sku")
                update_data["updated_at"] = now
                Product.objects.filter(pk=existing_product.pk).update(**update_data)
                product = Product.objects.get(pk=existing_product.pk)
                updated += 1
            else:
                product = Product(**product_data)
                product.save()
                created += 1
            save_imported_image(product, "primary_image", image_data)
            if product.is_company_grouped:
                product.options.all().delete()
                Product.objects.filter(pk=product.pk).update(has_options=False, price=Decimal("0"), updated_at=now)
            for import_key in import_keys:
                product_lookup[import_key] = product

        option_products = set()
        for row in option_rows:
            product = resolve_imported_product(product_lookup, row["product_key"])
            if product is None:
                continue
            if product.is_company_grouped:
                continue
            option_products.add(product.pk)
            option_data = {
                "price": row["price"],
                "is_default": row["is_default"],
                "is_available": row["is_available"],
                "display_order": row["display_order"],
            }
            if row["is_default"]:
                product.options.update(is_default=False)
            option = product.options.filter(name__iexact=row["name"]).first()
            if option:
                ProductOption.objects.filter(pk=option.pk).update(name=row["name"], **option_data)
            else:
                ProductOption.objects.create(product=product, name=row["name"], **option_data)

        if option_products:
            Product.objects.filter(pk__in=option_products).update(
                has_options=True,
                product_type=Product.PRODUCT_TYPE_NORMAL,
                updated_at=now,
            )

        company_lookup = {}
        for row in company_rows:
            company = Company.objects.filter(code__iexact=row["company_id"]).first()
            company_data = {
                "code": row["company_id"],
                "name": row["name"],
                "external_logo_url": row["external_logo_url"],
                "display_order": row["order"],
                "is_active": row["is_active"],
            }
            if company is None:
                company = Company.objects.create(**company_data)
            else:
                Company.objects.filter(pk=company.pk).update(**company_data, updated_at=now)
                company = Company.objects.get(pk=company.pk)
            merge_legacy_company_records(company)
            save_imported_image(company, "logo", row.get("logo"))
            company_lookup[row["company_id"].lower()] = company

        product_company_lookup = {}
        company_products = set()

        for row in company_option_rows:
            product = resolve_imported_product(product_lookup, row["product_key"])
            if product is None:
                continue
            company_id = row["company_id"].lower()
            company = company_lookup.get(company_id)
            if company is None:
                company = Company.objects.filter(code__iexact=row["company_id"]).first()
            if company is None:
                continue
            company_products.add(product.pk)
            company_key = (row["product_key"], company_id)
            product_company = product_company_lookup.get(company_key)
            if product_company is None:
                product_company, _ = ProductCompany.objects.get_or_create(
                    product=product,
                    company=company,
                )
                product_company_lookup[company_key] = product_company
            option_data = {
                "price": row["price"],
                "is_available": row["is_available"],
                "order": row["order"],
            }
            option = product_company.options.filter(name__iexact=row["name"]).first()
            if option:
                ProductCompanyOption.objects.filter(pk=option.pk).update(name=row["name"], **option_data)
            else:
                ProductCompanyOption.objects.create(
                    company=product_company,
                    name=row["name"],
                    **option_data,
                )

        if company_products:
            Product.objects.filter(pk__in=company_products).update(
                product_type=Product.PRODUCT_TYPE_COMPANY_GROUPED,
                has_options=False,
                price=Decimal("0"),
                updated_at=now,
            )
            ProductOption.objects.filter(product_id__in=company_products).delete()

    return created, updated


def validate_delivery_area_excel_workbook(excel_file, language=DEFAULT_LANGUAGE):
    dashboard_ui = get_dashboard_strings(language)
    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
    except Exception:
        return {}, [dashboard_ui["excel_read_error"]]

    areas_sheet = workbook["Areas"] if "Areas" in workbook.sheetnames else workbook.active
    area_header_positions, area_sheet_rows = build_sheet_rows(areas_sheet)
    missing_headers = [
        header for header in DELIVERY_AREA_EXCEL_REQUIRED_HEADERS if header not in area_header_positions
    ]
    if missing_headers:
        return {}, [f"Areas sheet missing required column: {', '.join(missing_headers)}"]

    valid_areas = []
    valid_sub_areas = []
    errors = []
    seen_area_names = set()
    known_area_names = {area.name.strip().lower() for area in DeliveryArea.objects.all()}
    sub_header_positions = {}
    sub_sheet_rows = []
    sub_area_area_names = set()
    if "SubAreas" in workbook.sheetnames:
        sub_header_positions, sub_sheet_rows = build_sheet_rows(workbook["SubAreas"])
        sub_area_area_names = {
            normalize_excel_text(row.get("area_name")).lower()
            for _, row in sub_sheet_rows
            if normalize_excel_text(row.get("area_name"))
        }

    for row_number, row in area_sheet_rows:
        row_errors = []
        area_name = normalize_excel_text(row.get("area_name"))
        if not area_name:
            row_errors.append("area_name is required.")
        area_key = area_name.lower()
        if area_key in seen_area_names:
            row_errors.append("This area appears more than once in this file.")
        if area_key:
            seen_area_names.add(area_key)
            known_area_names.add(area_key)

        has_sub_areas, boolean_error = parse_excel_boolean(row.get("has_sub_areas"), False)
        if boolean_error:
            row_errors.append(f"has_sub_areas: {boolean_error}")
        has_sub_areas = has_sub_areas or bool(area_key and area_key in sub_area_area_names)

        delivery_fee, fee_error = parse_excel_decimal(row.get("delivery_fee"))
        if fee_error and has_sub_areas and row.get("delivery_fee") in (None, ""):
            delivery_fee = Decimal("0")
        elif fee_error:
            row_errors.append(f"delivery_fee: {fee_error}")

        display_order, order_error = parse_excel_integer(row.get("display_order"), 0)
        if order_error:
            row_errors.append(f"display_order: {order_error}")

        is_active, active_error = parse_excel_boolean(row.get("is_active"), True)
        if active_error:
            row_errors.append(f"is_active: {active_error}")

        if row_errors:
            errors.append(f"Areas row {row_number}: {' '.join(row_errors)}")
            continue

        valid_areas.append(
            {
                "name": area_name,
                "has_sub_areas": has_sub_areas,
                "delivery_fee": delivery_fee,
                "display_order": display_order,
                "is_active": is_active,
            }
        )

    if "SubAreas" in workbook.sheetnames:
        if sub_sheet_rows:
            missing_sub_headers = [
                header for header in DELIVERY_SUB_AREA_EXCEL_REQUIRED_HEADERS if header not in sub_header_positions
            ]
            if missing_sub_headers:
                errors.append(f"SubAreas sheet missing required column: {', '.join(missing_sub_headers)}")

        seen_sub_area_keys = set()
        for row_number, row in sub_sheet_rows:
            row_errors = []
            area_name = normalize_excel_text(row.get("area_name"))
            sub_area_name = normalize_excel_text(row.get("sub_area_name"))
            area_key = area_name.lower()

            if not area_name:
                row_errors.append("area_name is required.")
            elif area_key not in known_area_names:
                row_errors.append(f"area_name '{area_name}' does not match an imported or existing area.")
            if not sub_area_name:
                row_errors.append("sub_area_name is required.")

            delivery_fee, fee_error = parse_excel_decimal(row.get("delivery_fee"))
            if fee_error:
                row_errors.append(f"delivery_fee: {fee_error}")

            display_order, order_error = parse_excel_integer(row.get("display_order"), 0)
            if order_error:
                row_errors.append(f"display_order: {order_error}")

            is_active, active_error = parse_excel_boolean(row.get("is_active"), True)
            if active_error:
                row_errors.append(f"is_active: {active_error}")

            sub_key = (area_key, sub_area_name.lower())
            if area_key and sub_area_name:
                if sub_key in seen_sub_area_keys:
                    row_errors.append("This sub-area appears more than once for the same area in this file.")
                seen_sub_area_keys.add(sub_key)

            if row_errors:
                errors.append(f"SubAreas row {row_number}: {' '.join(row_errors)}")
                continue

            valid_sub_areas.append(
                {
                    "area_name": area_name,
                    "name": sub_area_name,
                    "delivery_fee": delivery_fee,
                    "display_order": display_order,
                    "is_active": is_active,
                }
            )

    if not valid_areas and not errors:
        errors.append("The uploaded workbook has no delivery area rows.")

    return {"areas": valid_areas, "sub_areas": valid_sub_areas}, errors


def import_delivery_area_excel_rows(valid_rows):
    area_rows = valid_rows.get("areas", [])
    sub_area_rows = valid_rows.get("sub_areas", [])
    created = 0
    updated = 0
    area_lookup = {}
    sub_area_area_names = {row["area_name"].strip().lower() for row in sub_area_rows}

    with transaction.atomic():
        for row in area_rows:
            area_key = row["name"].strip().lower()
            row_data = row.copy()
            if area_key in sub_area_area_names:
                row_data["has_sub_areas"] = True
                row_data["delivery_fee"] = Decimal("0")
            area = DeliveryArea.objects.filter(name__iexact=row["name"]).first()
            if area:
                for field, value in row_data.items():
                    setattr(area, field, value)
                area.save()
                updated += 1
            else:
                area = DeliveryArea.objects.create(**row_data)
                created += 1
            area_lookup[area_key] = area

        for row in sub_area_rows:
            area_key = row["area_name"].strip().lower()
            area = area_lookup.get(area_key) or DeliveryArea.objects.filter(name__iexact=row["area_name"]).first()
            if area is None:
                continue
            area.has_sub_areas = True
            area.delivery_fee = Decimal("0")
            area.save(update_fields=["has_sub_areas", "delivery_fee", "updated_at"])
            sub_area = area.sub_areas.filter(name__iexact=row["name"]).first()
            sub_data = {
                "delivery_fee": row["delivery_fee"],
                "display_order": row["display_order"],
                "is_active": row["is_active"],
            }
            if sub_area:
                DeliverySubArea.objects.filter(pk=sub_area.pk).update(name=row["name"], **sub_data)
            else:
                DeliverySubArea.objects.create(area=area, name=row["name"], **sub_data)

    return created, updated


def product_options_have_rows(option_formset):
    return any(
        form.cleaned_data
        and not form.cleaned_data.get("DELETE")
        and form.cleaned_data.get("name")
        and form.cleaned_data.get("price") is not None
        for form in option_formset.forms
    )


def product_company_form_has_row(company_form):
    cleaned_data = getattr(company_form, "cleaned_data", None) or {}
    return bool(cleaned_data and not cleaned_data.get("DELETE") and cleaned_data.get("company"))


def product_company_options_have_rows(option_formset):
    return any(
        form.cleaned_data
        and not form.cleaned_data.get("DELETE")
        and form.cleaned_data.get("name")
        and form.cleaned_data.get("price") is not None
        for form in option_formset.forms
    )


def product_companies_have_rows(company_formset, option_formsets):
    return any(
        product_company_form_has_row(company_form) and product_company_options_have_rows(option_formsets[index])
        for index, company_form in enumerate(company_formset.forms)
    )


def validate_product_company_formsets(company_formset, option_formsets, dashboard_ui):
    has_company_with_options = False
    is_valid = True
    for index, company_form in enumerate(company_formset.forms):
        if not product_company_form_has_row(company_form):
            continue
        option_formset = option_formsets[index]
        if product_company_options_have_rows(option_formset):
            has_company_with_options = True
            continue
        option_formset._non_form_errors = option_formset.error_class(
            [dashboard_ui["company_options_required"]]
        )
        is_valid = False
    return has_company_with_options, is_valid


def save_product_options(product, option_formset):
    if not product.has_options:
        product.options.all().delete()
        return
    option_formset.instance = product
    option_formset.save()


def save_product_companies(product, company_formset, company_option_formsets):
    if not product.is_company_grouped:
        product.companies.all().delete()
        return

    company_formset.instance = product
    for index, company_form in enumerate(company_formset.forms):
        cleaned_data = getattr(company_form, "cleaned_data", None) or {}
        if not cleaned_data:
            continue
        if cleaned_data.get("DELETE"):
            if company_form.instance.pk:
                company_form.instance.delete()
            continue
        if not cleaned_data.get("company"):
            continue

        product_company = company_form.save(commit=False)
        product_company.product = product
        product_company.save()
        option_formset = company_option_formsets[index]
        option_formset.instance = product_company
        option_formset.save()


def build_company_option_formsets(company_formset, data=None):
    return [
        ProductCompanyOptionFormSet(
            data=data,
            instance=company_form.instance,
            prefix=f"company-{index}-options",
        )
        for index, company_form in enumerate(company_formset.forms)
    ]


def build_empty_company_option_formset():
    return ProductCompanyOptionFormSet(prefix="company-__company_index__-options")


def build_company_blocks(company_formset, company_option_formsets):
    return [
        {
            "company_form": company_form,
            "option_formset": company_option_formsets[index],
            "index": index,
        }
        for index, company_form in enumerate(company_formset.forms)
    ]


class ManagerLoginView(DashboardLocalizationMixin, LoginView):
    template_name = "dashboard/login.html"
    authentication_form = ManagerLoginForm
    redirect_authenticated_user = True

    def form_valid(self, form):
        user = form.get_user()
        if not user.is_staff:
            form.add_error(None, self.get_dashboard_ui()["staff_only_error"])
            return self.form_invalid(form)
        return super().form_valid(form)


class ManagerLogoutView(LogoutView):
    next_page = reverse_lazy("core:home")


class DashboardHomeView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/index.html"

    def get_context_data(self, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        context = super().get_context_data(**kwargs)
        context["category_count"] = Category.objects.count()
        context["product_count"] = Product.objects.count()
        context["available_product_count"] = Product.objects.filter(is_available=True).count()
        context["promotion_count"] = Promotion.objects.filter(is_active=True).count()
        context["delivery_area_count"] = DeliveryArea.objects.count()
        context["customer_count"] = CustomerProfile.objects.count()
        context["order_count"] = CustomerOrder.objects.count()
        context["pending_order_count"] = CustomerOrder.objects.filter(status__in=CustomerOrder.ACTIVE_STATUSES).count()
        return context


class CenterStatusManageView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/center_status.html"

    def get_center_status(self):
        return CenterStatus.get_current().refresh_availability()

    def get_context_data(self, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        context = super().get_context_data(**kwargs)
        center_status = self.get_center_status()
        remaining_seconds = max(int(center_status.remaining_busy_time().total_seconds()), 0)
        remaining_minutes = (remaining_seconds + 59) // 60
        context["center_status"] = center_status
        context["remaining_seconds"] = remaining_seconds
        context["remaining_minutes"] = remaining_minutes
        context["form"] = kwargs.get("form") or CenterStatusForm(
            instance=center_status,
            language=self.get_language(),
        )
        return context

    def post(self, request, *args, **kwargs):
        center_status = self.get_center_status()
        form = CenterStatusForm(request.POST, instance=center_status, language=self.get_language())
        if form.is_valid():
            form.save(user=request.user)
            sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
            messages.success(request, self.get_dashboard_ui()["center_status_updated"])
            return redirect("dashboard:center-status")
        return self.render_to_response(self.get_context_data(form=form))


CUSTOMER_ACCESS_SESSION_KEY = "dashboard_customer_access_granted"


def attach_dashboard_order_display(orders, language):
    return attach_orders_display(
        orders,
        get_dashboard_strings(language),
        language,
        print_auto_accepted_order=True,
    )


def parse_report_date(value):
    if not value:
        return None, None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date(), None
    except ValueError:
        return None, "Use YYYY-MM-DD."


def get_sales_report_filters(request, dashboard_ui, require_dates=False):
    single_day = (request.GET.get("single_day") or "").strip()
    from_value = (request.GET.get("from_date") or "").strip()
    to_value = (request.GET.get("to_date") or "").strip()
    if single_day:
        from_value = single_day
        to_value = single_day

    errors = []
    from_date, from_error = parse_report_date(from_value)
    to_date, to_error = parse_report_date(to_value)
    if from_error:
        errors.append(f"{dashboard_ui['sales_from_date']}: {from_error}")
    if to_error:
        errors.append(f"{dashboard_ui['sales_to_date']}: {to_error}")
    if require_dates and not from_value:
        errors.append(dashboard_ui["sales_from_date_required"])
    if require_dates and not to_value:
        errors.append(dashboard_ui["sales_to_date_required"])
    if from_date and to_date and from_date > to_date:
        errors.append(dashboard_ui["sales_date_order_error"])

    return {
        "from_date": from_value,
        "to_date": to_value,
        "single_day": single_day,
        "from_date_obj": from_date,
        "to_date_obj": to_date,
        "errors": errors,
    }


def get_sales_report_orders(from_date, to_date):
    start_at = timezone.make_aware(datetime.combine(from_date, time.min))
    end_at = timezone.make_aware(datetime.combine(to_date, time.max))
    return CustomerOrder.objects.filter(
        status__in=SALES_REPORT_STATUSES,
        created_at__gte=start_at,
        created_at__lte=end_at,
    )


def get_order_day_filter(request, dashboard_ui):
    day_value = (request.GET.get("order_day") or "").strip()
    day, day_error = parse_report_date(day_value)
    errors = []
    if day_error:
        errors.append(f"{dashboard_ui['orders_filter_day']}: {day_error}")
    return {"day": day_value, "day_obj": day, "errors": errors}


def filter_orders_by_day(queryset, day):
    if day is None:
        return queryset
    start_at = timezone.make_aware(datetime.combine(day, time.min))
    end_at = timezone.make_aware(datetime.combine(day, time.max))
    return queryset.filter(created_at__gte=start_at, created_at__lte=end_at)


def build_sales_report(from_date, to_date, language):
    orders = get_sales_report_orders(from_date, to_date)
    item_rows = (
        CustomerOrderItem.objects.filter(order__in=orders)
        .values("title", "category_label", "company_label", "selected_option_label", "unit_price")
        .annotate(quantity_sold=Sum("quantity"), total_product_price=Sum("line_total_max"))
        .order_by("category_label", "title", "company_label", "selected_option_label", "unit_price")
    )
    rows = []
    for row in item_rows:
        unit_price = row["unit_price"] or Decimal("0")
        total_product_price = row["total_product_price"] or Decimal("0")
        rows.append(
            {
                **row,
                "title": " - ".join(
                    part
                    for part in (
                        row["title"],
                        row.get("company_label"),
                        row.get("selected_option_label"),
                    )
                    if part
                ),
                "display_unit_price": format_syp(unit_price, language),
                "display_total_product_price": format_syp(total_product_price, language),
            }
        )

    summary = orders.aggregate(grand_total=Sum("total_max"), delivery_total=Sum("delivery_fee"))
    order_count = orders.count()
    grand_total = summary["grand_total"] or Decimal("0")
    delivery_total = summary["delivery_total"] or Decimal("0")
    return {
        "rows": rows,
        "summary": {
            "order_count": order_count,
            "grand_total": grand_total,
            "delivery_total": delivery_total,
            "display_grand_total": format_syp(grand_total, language),
            "display_delivery_total": format_syp(delivery_total, language),
            "from_date": from_date,
            "to_date": to_date,
        },
    }


def decimal_to_excel_number(value):
    return int(value or 0)


class CustomerAccessRequiredMixin:
    def has_customer_access(self):
        return bool(self.request.session.get(CUSTOMER_ACCESS_SESSION_KEY))


class CustomerListView(CustomerAccessRequiredMixin, DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/customers.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["has_customer_access"] = self.has_customer_access()
        context["access_form"] = kwargs.get("access_form") or CustomerAccessPasswordForm(language=self.get_language())
        if context["has_customer_access"]:
            search_query = self.request.GET.get("q", "").strip()
            customers = CustomerProfile.objects.select_related("user").order_by("full_name", "phone_number")
            if search_query:
                customers = customers.filter(
                    Q(full_name__icontains=search_query)
                    | Q(phone_number__icontains=search_query)
                    | Q(user__username__icontains=search_query)
                )
            context["customers"] = customers
            context["search_query"] = search_query
        return context

    def post(self, request, *args, **kwargs):
        form = CustomerAccessPasswordForm(request.POST, language=self.get_language())
        if form.is_valid():
            request.session[CUSTOMER_ACCESS_SESSION_KEY] = True
            request.session.modified = True
            messages.success(request, self.get_dashboard_ui()["customers_unlocked"])
            return redirect("dashboard:customers")
        return self.render_to_response(self.get_context_data(access_form=form))


class CustomerDetailView(CustomerAccessRequiredMixin, DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/customer_detail.html"

    def dispatch(self, request, *args, **kwargs):
        if not self.has_customer_access():
            return redirect("dashboard:customers")
        self.customer = get_object_or_404(
            CustomerProfile.objects.select_related("user").prefetch_related("addresses__area", "addresses__sub_area"),
            pk=kwargs["pk"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        language = self.get_language()
        orders = list(self.customer.orders.prefetch_related("items").all())
        context["customer"] = self.customer
        context["password_form"] = kwargs.get("password_form") or CustomerPasswordChangeForm(language=language)
        context["orders"] = attach_dashboard_order_display(orders, language)
        return context

    def post(self, request, *args, **kwargs):
        form = CustomerPasswordChangeForm(request.POST, language=self.get_language())
        if form.is_valid():
            self.customer.user.set_password(form.cleaned_data["new_password"])
            self.customer.user.save(update_fields=["password"])
            messages.success(request, self.get_dashboard_ui()["customer_password_updated"])
            return redirect("dashboard:customer-detail", pk=self.customer.pk)
        return self.render_to_response(self.get_context_data(password_form=form))


class DashboardOrdersView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/orders.html"

    def get_context_data(self, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        context = super().get_context_data(**kwargs)
        dashboard_ui = self.get_dashboard_ui()
        report_filters = get_sales_report_filters(self.request, dashboard_ui)
        order_day_filter = get_order_day_filter(self.request, dashboard_ui)
        report = None
        if report_filters["from_date_obj"] and report_filters["to_date_obj"] and not report_filters["errors"]:
            report = build_sales_report(
                report_filters["from_date_obj"],
                report_filters["to_date_obj"],
                self.get_language(),
            )
        orders_queryset = (
            CustomerOrder.objects.select_related("profile", "address").prefetch_related("items").all()
        )
        if not order_day_filter["errors"]:
            orders_queryset = filter_orders_by_day(orders_queryset, order_day_filter["day_obj"])
        orders = list(orders_queryset)
        context["orders"] = attach_dashboard_order_display(orders, self.get_language())
        context["order_day_filter"] = order_day_filter
        context["order_filter_errors"] = order_day_filter["errors"]
        context["sales_report_filters"] = report_filters
        context["sales_report"] = report
        context["sales_report_errors"] = report_filters["errors"]
        suggested_minutes = suggested_expected_minutes()
        for order in context["orders"]:
            order.suggested_expected_minutes = suggested_minutes
        return context


class SalesReportExcelExportView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        dashboard_ui = self.get_dashboard_ui()
        report_filters = get_sales_report_filters(request, dashboard_ui, require_dates=True)
        if report_filters["errors"]:
            for error in report_filters["errors"]:
                messages.error(request, error)
            return redirect("dashboard:orders")

        report = build_sales_report(
            report_filters["from_date_obj"],
            report_filters["to_date_obj"],
            self.get_language(),
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sales Report"
        worksheet.append([dashboard_ui["sales_report"]])
        worksheet.append(
            [
                dashboard_ui["sales_date_range"],
                f"{report_filters['from_date']} - {report_filters['to_date']}",
            ]
        )
        worksheet.append([])
        worksheet.append(
            [
                dashboard_ui["sales_product_name"],
                dashboard_ui["sales_category"],
                dashboard_ui["sales_quantity_sold"],
                dashboard_ui["sales_unit_price"],
                dashboard_ui["sales_total_product_price"],
            ]
        )
        for row in report["rows"]:
            worksheet.append(
                [
                    row["title"],
                    row["category_label"],
                    row["quantity_sold"] or 0,
                    decimal_to_excel_number(row["unit_price"]),
                    decimal_to_excel_number(row["total_product_price"]),
                ]
            )
        worksheet.append([])
        worksheet.append([dashboard_ui["sales_grand_total"], decimal_to_excel_number(report["summary"]["grand_total"])])
        worksheet.append([dashboard_ui["sales_delivery_total"], decimal_to_excel_number(report["summary"]["delivery_total"])])
        worksheet.append([dashboard_ui["sales_order_count"], report["summary"]["order_count"]])
        worksheet.append(
            [
                dashboard_ui["sales_date_range"],
                f"{report_filters['from_date']} - {report_filters['to_date']}",
            ]
        )
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(len(header) + 6, 18)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"sales-report-{report_filters['from_date']}-to-{report_filters['to_date']}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PendingOrdersView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/pending_orders.html"

    def get_context_data(self, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        context = super().get_context_data(**kwargs)
        orders = list(
            CustomerOrder.objects.select_related("profile", "address")
            .prefetch_related("items")
            .filter(status__in=CustomerOrder.ACTIVE_STATUSES)
            .order_by("created_at", "id")
        )
        context["orders"] = attach_dashboard_order_display(orders, self.get_language())
        return context


class DashboardOrderDetailView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/order_detail.html"

    def dispatch(self, request, *args, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        self.order = get_object_or_404(
            CustomerOrder.objects.select_related("profile", "address").prefetch_related("items"),
            pk=kwargs["pk"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        language = self.get_language()
        context["order"] = attach_order_display(
            self.order,
            self.get_dashboard_ui(),
            language,
            print_auto_accepted_order=True,
        )
        context["accept_form"] = kwargs.get("accept_form") or OrderAcceptForm(
            language=language,
            suggested_minutes=suggested_expected_minutes(),
        )
        return context


class DashboardOrderAcceptView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        order = get_object_or_404(CustomerOrder, pk=pk, status=CustomerOrder.STATUS_WAITING_ACCEPT)
        form = OrderAcceptForm(
            request.POST,
            language=self.get_language(),
            suggested_minutes=suggested_expected_minutes(),
        )
        if not form.is_valid():
            detail = DashboardOrderDetailView()
            detail.setup(request, pk=pk)
            detail.order = order
            return detail.render_to_response(detail.get_context_data(accept_form=form))

        mark_order_accepted(
            order,
            expected_time_minutes=form.cleaned_data["expected_time_minutes"],
            print_invoice=True,
        )
        messages.success(request, self.get_dashboard_ui()["order_accepted"])
        return redirect("dashboard:pending-order-detail", pk=order.pk)


class DashboardOrderAdvanceView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        sync_center_status_and_auto_accept_waiting_orders(print_invoices=True)
        order = get_object_or_404(CustomerOrder, pk=pk)
        if order.status in {CustomerOrder.STATUS_WAITING_ACCEPT, CustomerOrder.STATUS_WAITING_BUSY_CENTER}:
            messages.error(request, self.get_dashboard_ui()["invalid_order_status_change"])
            return redirect("dashboard:pending-order-detail", pk=order.pk)
        next_status = order.get_next_status()
        requested_status = request.POST.get("status")
        if not next_status or requested_status != next_status:
            messages.error(request, self.get_dashboard_ui()["invalid_order_status_change"])
            return redirect("dashboard:pending-order-detail", pk=order.pk)

        order.status = next_status
        update_fields = ["status", "updated_at"]
        if next_status == CustomerOrder.STATUS_DONE:
            order.completed_at = timezone.now()
            update_fields.append("completed_at")
        order.save(update_fields=update_fields)
        messages.success(request, self.get_dashboard_ui()["order_status_updated"])
        if order.status == CustomerOrder.STATUS_DONE:
            return redirect("dashboard:orders")
        return redirect("dashboard:pending-order-detail", pk=order.pk)


class CategoryManageView(DashboardLocalizationMixin, StaffRequiredMixin, CreateView):
    template_name = "dashboard/manage_categories.html"
    form_class = CategoryForm
    success_url = reverse_lazy("dashboard:categories")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        language = self.get_language()
        search_query = self.request.GET.get("q", "").strip()
        items = Category.objects.select_related("parent").order_by(
            "parent__display_order",
            "parent__name",
            "display_order",
            "name",
        )
        if search_query:
            items = items.filter(
                Q(name__icontains=search_query)
                | Q(name_ar__icontains=search_query)
                | Q(description__icontains=search_query)
                | Q(description_ar__icontains=search_query)
                | Q(slug__icontains=search_query)
            )
        localized_items = localize_queryset(
            items,
            language,
            ["name", "description"],
        )
        for item in localized_items:
            localize_instance(item.parent, language, ["name", "description"])
        context["items"] = localized_items
        context["search_query"] = search_query
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["category_saved"])
        return super().form_valid(form)


class CategoryUpdateView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = "dashboard/form_page.html"
    success_url = reverse_lazy("dashboard:categories")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["edit_category"]
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["category_updated"])
        return super().form_valid(form)


class CategoryDeleteView(DashboardLocalizationMixin, StaffRequiredMixin, DeleteView):
    model = Category
    success_url = reverse_lazy("dashboard:categories")

    def post(self, request, *args, **kwargs):
        messages.success(request, self.get_dashboard_ui()["category_deleted"])
        return super().post(request, *args, **kwargs)


class ProductManageView(DashboardLocalizationMixin, StaffRequiredMixin, CreateView):
    template_name = "dashboard/manage_products.html"
    form_class = ProductForm
    success_url = reverse_lazy("dashboard:products")

    def get_option_formset(self, data=None):
        return ProductOptionFormSet(data=data, prefix="options")

    def get_company_formset(self, data=None, files=None):
        return ProductCompanyFormSet(data=data, files=files, prefix="companies")

    def get_company_option_formsets(self, company_formset, data=None):
        return build_company_option_formsets(company_formset, data=data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(get_product_manage_context(self.get_language(), product_form=context.get("form")))
        context["option_formset"] = kwargs.get("option_formset") or self.get_option_formset()
        context["company_formset"] = kwargs.get("company_formset") or self.get_company_formset()
        context["company_option_formsets"] = kwargs.get("company_option_formsets") or self.get_company_option_formsets(
            context["company_formset"]
        )
        context["company_blocks"] = build_company_blocks(
            context["company_formset"],
            context["company_option_formsets"],
        )
        context["empty_company_option_formset"] = build_empty_company_option_formset()
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        option_formset = self.get_option_formset(request.POST)
        company_formset = self.get_company_formset(request.POST, request.FILES)
        company_option_formsets = self.get_company_option_formsets(company_formset, request.POST)
        if form.is_valid():
            is_company_grouped = form.cleaned_data.get("product_type") == Product.PRODUCT_TYPE_COMPANY_GROUPED
            normal_options_valid = True if is_company_grouped else option_formset.is_valid()
            company_formset_valid = company_formset.is_valid() if is_company_grouped else True
            company_options_valid = (
                all(option_formset.is_valid() for option_formset in company_option_formsets)
                if is_company_grouped
                else True
            )
            if not normal_options_valid or not company_formset_valid or not company_options_valid:
                return self.form_invalid(form, option_formset, company_formset, company_option_formsets)
            has_companies, company_options_valid = validate_product_company_formsets(
                company_formset,
                company_option_formsets,
                self.get_dashboard_ui(),
            )
            if is_company_grouped and not has_companies:
                company_formset._non_form_errors = company_formset.error_class(
                    [self.get_dashboard_ui()["product_companies_required"]]
                )
            elif is_company_grouped and not company_options_valid:
                pass
            elif form.cleaned_data.get("has_options") and not product_options_have_rows(option_formset):
                option_formset._non_form_errors = option_formset.error_class(
                    [self.get_dashboard_ui()["product_options_required"]]
                )
            else:
                return self.save_valid_product(form, option_formset, company_formset, company_option_formsets)
        return self.form_invalid(form, option_formset, company_formset, company_option_formsets)

    def save_valid_product(self, form, option_formset, company_formset, company_option_formsets):
        with transaction.atomic():
            product = form.save()
            save_product_options(product, option_formset)
            save_product_companies(product, company_formset, company_option_formsets)
        messages.success(self.request, self.get_dashboard_ui()["product_saved"])
        return redirect(self.success_url)

    def form_invalid(self, form, option_formset=None, company_formset=None, company_option_formsets=None):
        context = self.get_context_data(
            form=form,
            option_formset=option_formset,
            company_formset=company_formset,
            company_option_formsets=company_option_formsets,
        )
        return self.render_to_response(context)


class ProductListView(DashboardLocalizationMixin, StaffRequiredMixin, TemplateView):
    template_name = "dashboard/product_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get("q", "").strip()
        filters = {
            "name": self.request.GET.get("filter_name", ""),
            "category": self.request.GET.get("filter_category", ""),
            "pricing_type": self.request.GET.get("filter_pricing_type", ""),
            "status": self.request.GET.get("filter_status", ""),
        }
        context.update(get_product_list_context(self.get_language(), search_query=search_query, filters=filters))
        return context


class ProductUpdateView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = "dashboard/form_page.html"
    success_url = reverse_lazy("dashboard:product-list")

    def get_option_formset(self, data=None):
        return ProductOptionFormSet(data=data, instance=self.object, prefix="options")

    def get_company_formset(self, data=None, files=None):
        return ProductCompanyFormSet(data=data, files=files, instance=self.object, prefix="companies")

    def get_company_option_formsets(self, company_formset, data=None):
        return build_company_option_formsets(company_formset, data=data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["edit_product"]
        context["option_formset"] = kwargs.get("option_formset") or self.get_option_formset()
        context["company_formset"] = kwargs.get("company_formset") or self.get_company_formset()
        context["company_option_formsets"] = kwargs.get("company_option_formsets") or self.get_company_option_formsets(
            context["company_formset"]
        )
        context["company_blocks"] = build_company_blocks(
            context["company_formset"],
            context["company_option_formsets"],
        )
        context["empty_company_option_formset"] = build_empty_company_option_formset()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        option_formset = self.get_option_formset(request.POST)
        company_formset = self.get_company_formset(request.POST, request.FILES)
        company_option_formsets = self.get_company_option_formsets(company_formset, request.POST)
        if form.is_valid():
            is_company_grouped = form.cleaned_data.get("product_type") == Product.PRODUCT_TYPE_COMPANY_GROUPED
            normal_options_valid = True if is_company_grouped else option_formset.is_valid()
            company_formset_valid = company_formset.is_valid() if is_company_grouped else True
            company_options_valid = (
                all(option_formset.is_valid() for option_formset in company_option_formsets)
                if is_company_grouped
                else True
            )
            if not normal_options_valid or not company_formset_valid or not company_options_valid:
                return self.form_invalid(form, option_formset, company_formset, company_option_formsets)
            has_companies, company_options_valid = validate_product_company_formsets(
                company_formset,
                company_option_formsets,
                self.get_dashboard_ui(),
            )
            if is_company_grouped and not has_companies:
                company_formset._non_form_errors = company_formset.error_class(
                    [self.get_dashboard_ui()["product_companies_required"]]
                )
            elif is_company_grouped and not company_options_valid:
                pass
            elif form.cleaned_data.get("has_options") and not product_options_have_rows(option_formset):
                option_formset._non_form_errors = option_formset.error_class(
                    [self.get_dashboard_ui()["product_options_required"]]
                )
            else:
                return self.save_valid_product(form, option_formset, company_formset, company_option_formsets)
        return self.form_invalid(form, option_formset, company_formset, company_option_formsets)

    def save_valid_product(self, form, option_formset, company_formset, company_option_formsets):
        with transaction.atomic():
            product = form.save()
            save_product_options(product, option_formset)
            save_product_companies(product, company_formset, company_option_formsets)
        messages.success(self.request, self.get_dashboard_ui()["product_updated"])
        return redirect(self.success_url)

    def form_invalid(self, form, option_formset=None, company_formset=None, company_option_formsets=None):
        context = self.get_context_data(
            form=form,
            option_formset=option_formset,
            company_formset=company_formset,
            company_option_formsets=company_option_formsets,
        )
        return self.render_to_response(context)


class ProductDeleteView(DashboardLocalizationMixin, StaffRequiredMixin, DeleteView):
    model = Product
    success_url = reverse_lazy("dashboard:product-list")

    def post(self, request, *args, **kwargs):
        messages.success(request, self.get_dashboard_ui()["product_deleted"])
        return super().post(request, *args, **kwargs)


class ProductExcelTemplateView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        language = self.get_language()
        categories = list(
            Category.objects.select_related("parent")
            .filter(is_active=True)
            .order_by("parent__display_order", "parent__name", "display_order", "name")
        )
        sample_category = next((category for category in categories if category.parent_id), None)
        sample_category = sample_category or (categories[0] if categories else None)
        sample_category_value = sample_category.slug if sample_category is not None else "subcategory-slug"

        workbook = Workbook()
        products_sheet = workbook.active
        products_sheet.title = "Products"
        products_sheet.append(PRODUCT_EXCEL_HEADERS)
        products_sheet.append(
            [
                "Sample Product",
                sample_category_value,
                "Short product description",
                "Full product description",
                "",
                "normal",
                "inherit",
                "false",
                "custom",
                "false",
                "true",
                "piece",
                "true",
                "false",
                "https://example.com/product.jpg",
                "",
                "SAMPLE-001",
            ]
        )
        products_sheet.append(
            [
                "Running Shoes",
                sample_category_value,
                "Company grouped sample",
                "Choose the brand, then the option.",
                "",
                "company_grouped",
                "inherit",
                "false",
                "custom",
                "false",
                "false",
                "piece",
                "true",
                "false",
                "https://example.com/shoes.jpg",
                "",
                "SHOES-001",
            ]
        )
        options_sheet = workbook.create_sheet("Options")
        options_sheet.append(PRODUCT_OPTION_EXCEL_HEADERS)
        options_sheet.append(["SAMPLE-001", "Option 1", "1.08", "true", "true", "0"])
        options_sheet.append(["SAMPLE-001", "Option 2", "2.48", "false", "false", "1"])

        companies_sheet = workbook.create_sheet("Companies")
        companies_sheet.append(PRODUCT_COMPANY_EXCEL_HEADERS)
        companies_sheet.append(
            ["adidas", "Adidas", "https://example.com/adidas-logo.png", "", "0", "true"]
        )
        companies_sheet.append(
            ["nike", "Nike", "https://example.com/nike-logo.png", "", "1", "true"]
        )

        company_options_sheet = workbook.create_sheet("CompanyOptions")
        company_options_sheet.append(PRODUCT_COMPANY_OPTION_EXCEL_HEADERS)
        company_options_sheet.append(["SHOES-001", "adidas", "Size 42 Black", "15.50", "true", "0"])
        company_options_sheet.append(["SHOES-001", "adidas", "Size 43 White", "16.00", "true", "1"])
        company_options_sheet.append(["SHOES-001", "nike", "Size 42 Blue", "18.00", "true", "0"])

        categories_sheet = workbook.create_sheet("Categories")
        categories_sheet.append(CATEGORY_IMAGE_EXCEL_HEADERS)
        for category in categories:
            categories_sheet.append(
                [
                    category.slug,
                    "subcategory" if category.parent_id else "parent",
                    format_category_path(category.parent, language) if category.parent_id else "",
                    format_category_path(category, language),
                    category.slug,
                    category.cover_image.name if category.cover_image else "",
                    category.external_image_url,
                ]
            )

        help_sheet = workbook.create_sheet("Help")
        help_sheet.append(["Products sheet"])
        help_sheet.append(["Use the category column for the category or subcategory slug."])
        help_sheet.append(["For products under a parent section, use the subcategory slug, not the parent slug."])
        help_sheet.append(["product_type accepts normal or company_grouped. Leave blank for normal."])
        help_sheet.append(["Use image for a local image path or image URL saved into the primary image field."])
        help_sheet.append(["Use external_image_url only when the storefront should keep using an external URL."])
        help_sheet.append(["You can copy valid values from the Categories sheet."])
        help_sheet.append(["Options sheet"])
        help_sheet.append(["Use product_sku to connect normal product options to a product row in the Products sheet."])
        help_sheet.append(["Companies sheet"])
        help_sheet.append(["Use company_id and company_name to define reusable companies. Do not add product_sku here."])
        help_sheet.append(["Use logo for a local image path or image URL saved into the logo field."])
        help_sheet.append(["Use external_logo_url only when the storefront should keep using an external URL."])
        help_sheet.append(["CompanyOptions sheet"])
        help_sheet.append(["Use product_sku, company_id, and option_name for variants under a company."])
        help_sheet.append(["Categories sheet"])
        help_sheet.append(["For an existing category row, image updates its cover image before products import."])
        help_sheet.append(["Dollar-linked prices can use decimals such as 1.08 or 2.48."])

        header_fill = PatternFill("solid", fgColor="DDEBFF")
        header_font = Font(bold=True, color="0B2E63")
        for worksheet in (products_sheet, options_sheet, companies_sheet, company_options_sheet, categories_sheet, help_sheet):
            for header_cell in worksheet[1]:
                header_cell.fill = header_fill
                header_cell.font = header_font

        for worksheet in (products_sheet, options_sheet, companies_sheet, company_options_sheet, categories_sheet, help_sheet):
            for column_cells in worksheet.columns:
                header = column_cells[0].value or ""
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                    max(max_length + 4, len(header) + 4, 16),
                    44,
                )

        products_sheet.freeze_panes = "A2"
        options_sheet.freeze_panes = "A2"
        companies_sheet.freeze_panes = "A2"
        company_options_sheet.freeze_panes = "A2"
        categories_sheet.freeze_panes = "A2"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products-upload-template.xlsx"'
        return response


class ProductExcelUploadView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    template_name = "dashboard/manage_products.html"

    def post(self, request, *args, **kwargs):
        language = self.get_language()
        dashboard_ui = get_dashboard_strings(language)
        redirect_target = request.POST.get("next") or "dashboard:products"
        if redirect_target not in {"dashboard:products", "dashboard:product-list"}:
            redirect_target = "dashboard:products"
        upload_form = ProductExcelUploadForm(request.POST, request.FILES, language=language)
        if not upload_form.is_valid():
            return self.render_upload_response(request, language, redirect_target, upload_form=upload_form)

        merge_duplicate_products_for_import()
        valid_rows, excel_errors = validate_product_excel_workbook(
            upload_form.cleaned_data["excel_file"],
            language,
        )
        if excel_errors:
            messages.error(
                request,
                dashboard_ui["excel_import_complete"].format(
                    created=0,
                    updated=0,
                    failed=len(excel_errors),
                ),
            )
            return self.render_upload_response(
                request,
                language,
                redirect_target,
                upload_form=upload_form,
                excel_errors=excel_errors,
            )

        created, updated = import_product_excel_rows(valid_rows)
        messages.success(
            request,
            dashboard_ui["excel_import_complete"].format(created=created, updated=updated, failed=0),
        )
        return redirect(redirect_target)

    def render_upload_response(self, request, language, redirect_target, upload_form=None, excel_errors=None):
        if redirect_target == "dashboard:product-list":
            context = get_product_list_context(language, upload_form=upload_form, excel_errors=excel_errors)
            return render(request, "dashboard/product_list.html", context)
        context = get_product_manage_context(language, upload_form=upload_form, excel_errors=excel_errors)
        return render(request, self.template_name, context)


class PromotionManageView(DashboardLocalizationMixin, StaffRequiredMixin, CreateView):
    template_name = "dashboard/manage_promotions.html"
    form_class = PromotionForm
    success_url = reverse_lazy("dashboard:promotions")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        language = self.get_language()
        site_settings = get_active_site_settings() or get_dashboard_site_settings()
        items = localize_queryset(
            Promotion.objects.all(),
            language,
            ["title", "subtitle", "description", "badge_text", "cta_text"],
        )
        for item in items:
            set_promotion_display_price(item, language, site_settings)
        context["items"] = items
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["promotion_saved"])
        return super().form_valid(form)


class PromotionUpdateView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    model = Promotion
    form_class = PromotionForm
    template_name = "dashboard/form_page.html"
    success_url = reverse_lazy("dashboard:promotions")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["edit_promotion"]
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["promotion_updated"])
        return super().form_valid(form)


class PromotionDeleteView(DashboardLocalizationMixin, StaffRequiredMixin, DeleteView):
    model = Promotion
    success_url = reverse_lazy("dashboard:promotions")

    def post(self, request, *args, **kwargs):
        messages.success(request, self.get_dashboard_ui()["promotion_deleted"])
        return super().post(request, *args, **kwargs)


class DeliveryAreaManageView(DashboardLocalizationMixin, StaffRequiredMixin, CreateView):
    template_name = "dashboard/manage_delivery_areas.html"
    form_class = DeliveryAreaForm
    success_url = reverse_lazy("dashboard:delivery-areas")

    def get_sub_area_formset(self, data=None, instance=None):
        return DeliverySubAreaFormSet(
            data=data,
            instance=instance,
            prefix="subareas",
            form_kwargs={"language": self.get_language()},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(get_delivery_area_manage_context(self.get_language(), area_form=context.get("form")))
        context.setdefault("sub_area_formset", self.get_sub_area_formset())
        return context

    def has_configured_sub_area(self, sub_area_formset):
        for sub_area_form in sub_area_formset.forms:
            cleaned_data = getattr(sub_area_form, "cleaned_data", None) or {}
            if cleaned_data.get("DELETE"):
                continue
            if (cleaned_data.get("name") or "").strip() and cleaned_data.get("is_active"):
                return True
        return False

    def validate_sub_area_pricing(self, form, sub_area_formset):
        if not form.cleaned_data.get("has_sub_areas"):
            return True
        if self.has_configured_sub_area(sub_area_formset):
            return True
        form.add_error("has_sub_areas", self.get_dashboard_ui()["sub_area_pricing_requires_active_sub_area"])
        return False

    def form_invalid(self, form):
        return self.render_to_response(
            self.get_context_data(form=form, sub_area_formset=self.get_sub_area_formset(data=self.request.POST))
        )

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if not form.is_valid():
            return self.form_invalid(form)

        sub_area_formset = self.get_sub_area_formset(data=request.POST, instance=form.instance)
        with transaction.atomic():
            if sub_area_formset.is_valid() and self.validate_sub_area_pricing(form, sub_area_formset):
                area = form.save()
                sub_area_formset.instance = area
                sub_area_formset.save()
                messages.success(self.request, self.get_dashboard_ui()["delivery_area_saved"])
                return redirect(self.success_url)
            transaction.set_rollback(True)

        return self.render_to_response(
            self.get_context_data(form=form, sub_area_formset=sub_area_formset)
        )


class DeliveryAreaExcelTemplateView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        areas_sheet = workbook.active
        areas_sheet.title = "Areas"
        areas_sheet.append(DELIVERY_AREA_EXCEL_HEADERS)
        areas_sheet.append(["Mezzeh", "true", "", "0", "true"])
        areas_sheet.append(["Malki", "false", "15000", "1", "true"])

        sub_areas_sheet = workbook.create_sheet("SubAreas")
        sub_areas_sheet.append(DELIVERY_SUB_AREA_EXCEL_HEADERS)
        sub_areas_sheet.append(["Mezzeh", "East Mezzeh", "12000", "0", "true"])
        sub_areas_sheet.append(["Mezzeh", "West Mezzeh", "15000", "1", "true"])

        for worksheet in (areas_sheet, sub_areas_sheet):
            for column_cells in worksheet.columns:
                header = column_cells[0].value or ""
                worksheet.column_dimensions[column_cells[0].column_letter].width = max(len(header) + 4, 16)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="delivery-areas-upload-template.xlsx"'
        return response


class DeliveryAreaExcelUploadView(DashboardLocalizationMixin, StaffRequiredMixin, View):
    template_name = "dashboard/manage_delivery_areas.html"

    def post(self, request, *args, **kwargs):
        language = self.get_language()
        dashboard_ui = get_dashboard_strings(language)
        upload_form = DeliveryAreaExcelUploadForm(request.POST, request.FILES, language=language)
        if not upload_form.is_valid():
            return self.render_upload_response(request, language, upload_form=upload_form)

        valid_rows, excel_errors = validate_delivery_area_excel_workbook(
            upload_form.cleaned_data["excel_file"],
            language,
        )
        if excel_errors:
            messages.error(
                request,
                dashboard_ui["delivery_area_excel_import_complete"].format(
                    created=0,
                    updated=0,
                    failed=len(excel_errors),
                ),
            )
            return self.render_upload_response(
                request,
                language,
                upload_form=upload_form,
                excel_errors=excel_errors,
            )

        created, updated = import_delivery_area_excel_rows(valid_rows)
        messages.success(
            request,
            dashboard_ui["delivery_area_excel_import_complete"].format(created=created, updated=updated, failed=0),
        )
        return redirect("dashboard:delivery-areas")

    def render_upload_response(self, request, language, upload_form=None, excel_errors=None):
        context = get_delivery_area_manage_context(language, upload_form=upload_form, excel_errors=excel_errors)
        context["sub_area_formset"] = DeliverySubAreaFormSet(prefix="subareas", form_kwargs={"language": language})
        return render(request, self.template_name, context)


class DeliveryAreaUpdateView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    model = DeliveryArea
    form_class = DeliveryAreaForm
    template_name = "dashboard/delivery_area_form.html"
    success_url = reverse_lazy("dashboard:delivery-areas")

    def get_sub_area_formset(self, data=None):
        return DeliverySubAreaFormSet(
            data=data,
            instance=self.object,
            prefix="subareas",
            form_kwargs={"language": self.get_language()},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["edit_delivery_area"]
        context.setdefault("sub_area_formset", self.get_sub_area_formset())
        return context

    def has_configured_sub_area(self, sub_area_formset):
        for sub_area_form in sub_area_formset.forms:
            cleaned_data = getattr(sub_area_form, "cleaned_data", None) or {}
            if cleaned_data.get("DELETE"):
                continue
            if (cleaned_data.get("name") or "").strip() and cleaned_data.get("is_active"):
                return True
        return False

    def validate_sub_area_pricing(self, form, sub_area_formset):
        if not form.cleaned_data.get("has_sub_areas"):
            return True
        if self.has_configured_sub_area(sub_area_formset):
            return True
        form.add_error("has_sub_areas", self.get_dashboard_ui()["sub_area_pricing_requires_active_sub_area"])
        return False

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        sub_area_formset = self.get_sub_area_formset(data=request.POST)
        if (
            not form.is_valid()
            or not sub_area_formset.is_valid()
            or not self.validate_sub_area_pricing(form, sub_area_formset)
        ):
            return self.render_to_response(
                self.get_context_data(form=form, sub_area_formset=sub_area_formset)
            )

        try:
            with transaction.atomic():
                form.save()
                sub_area_formset.save()
        except ProtectedError:
            messages.error(request, self.get_dashboard_ui()["delivery_sub_area_in_use"])
            return self.render_to_response(
                self.get_context_data(form=form, sub_area_formset=sub_area_formset)
            )

        messages.success(self.request, self.get_dashboard_ui()["delivery_area_updated"])
        return redirect(self.success_url)


class DeliveryAreaDeleteView(DashboardLocalizationMixin, StaffRequiredMixin, DeleteView):
    model = DeliveryArea
    success_url = reverse_lazy("dashboard:delivery-areas")

    def post(self, request, *args, **kwargs):
        try:
            response = super().post(request, *args, **kwargs)
        except ProtectedError:
            messages.error(request, self.get_dashboard_ui()["delivery_area_in_use"])
            return redirect("dashboard:delivery-areas")
        messages.success(request, self.get_dashboard_ui()["delivery_area_deleted"])
        return response


class SiteSettingsManageView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    template_name = "dashboard/form_page.html"
    form_class = SiteSettingsForm
    success_url = reverse_lazy("dashboard:settings")

    def get_object(self, queryset=None):
        return get_dashboard_site_settings()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["store_settings"]
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["settings_updated"])
        return super().form_valid(form)


class DollarPriceManageView(DashboardLocalizationMixin, StaffRequiredMixin, UpdateView):
    template_name = "dashboard/form_page.html"
    form_class = DollarPriceForm
    success_url = reverse_lazy("dashboard:dollar-price")

    def get_object(self, queryset=None):
        return get_dashboard_site_settings()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.get_dashboard_ui()["edit_dollar_price"]
        context["submit_label"] = self.get_dashboard_ui()["save_dollar_price"]
        return context

    def form_valid(self, form):
        messages.success(self.request, self.get_dashboard_ui()["dollar_price_updated"])
        return super().form_valid(form)
