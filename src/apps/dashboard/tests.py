from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook
from PIL import Image

from apps.catalog.models import (
    Category,
    Company,
    Product,
    ProductCompany,
    ProductCompanyOption,
    ProductOption,
)
from apps.core.models import CustomerOrder, CustomerProfile

from .views import (
    CATEGORY_IMAGE_EXCEL_HEADERS,
    PRODUCT_COMPANY_EXCEL_HEADERS,
    PRODUCT_COMPANY_OPTION_EXCEL_HEADERS,
    PRODUCT_EXCEL_HEADERS,
    import_product_excel_rows,
    load_excel_image,
    validate_product_excel_workbook,
)


class CompanyExcelImportTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name="Shoes", slug="shoes")

    def product_row(self, **overrides):
        row = {
            "category": self.category,
            "name": "Running Shoes",
            "short_description": "",
            "description": "",
            "price": Decimal("0"),
            "product_type": Product.PRODUCT_TYPE_COMPANY_GROUPED,
            "price_link_mode": Product.BEHAVIOR_INHERIT,
            "is_price_linked_to_dollar": False,
            "sold_by_weight_mode": Product.BEHAVIOR_INHERIT,
            "sold_by_weight": False,
            "has_options": False,
            "unit_label": "piece",
            "is_available": True,
            "is_featured": False,
            "external_image_url": "",
            "sku": "SHOES-001",
            "_import_keys": [("sku", "shoes-001")],
        }
        row.update(overrides)
        return row

    def company_row(self, company_id, name):
        return {
            "company_id": company_id,
            "name": name,
            "external_logo_url": "",
            "order": 0,
            "is_active": True,
        }

    def company_option_row(self, company_id, name, price=Decimal("15.50")):
        return {
            "product_key": ("sku", "shoes-001"),
            "company_id": company_id,
            "name": name,
            "price": price,
            "is_available": True,
            "order": 0,
        }

    def option_row(self, name, price=Decimal("15.50")):
        return {
            "product_key": ("sku", "shoes-001"),
            "name": name,
            "price": price,
            "is_default": False,
            "is_available": True,
            "display_order": 0,
        }

    def import_rows(self, logo_url, company_name="Company One", is_price_linked_to_dollar=None):
        return import_product_excel_rows(
            {
                "products": [
                    {
                        "category": self.category,
                        "name": "Running Shoes",
                        "short_description": "",
                        "description": "",
                        "price": Decimal("0"),
                        "product_type": Product.PRODUCT_TYPE_COMPANY_GROUPED,
                        "price_link_mode": Product.BEHAVIOR_INHERIT,
                        "is_price_linked_to_dollar": False,
                        "sold_by_weight_mode": Product.BEHAVIOR_INHERIT,
                        "sold_by_weight": False,
                        "has_options": False,
                        "unit_label": "piece",
                        "is_available": True,
                        "is_featured": False,
                        "external_image_url": "",
                        "sku": "SHOES-001",
                        "_import_keys": [("sku", "shoes-001")],
                    }
                ],
                "options": [],
                "companies": [
                    {
                        "company_id": "Com-1",
                        "name": company_name,
                        "external_logo_url": logo_url,
                        "order": 2,
                        "is_active": True,
                    }
                ],
                "company_options": [
                    {
                        "product_key": ("sku", "shoes-001"),
                        "company_id": "Com-1",
                        "name": "Size 42",
                        "price": Decimal("15.50"),
                        "is_price_linked_to_dollar": is_price_linked_to_dollar,
                        "is_available": True,
                        "order": 0,
                    }
                ],
            }
        )

    def test_import_removes_product_companies_missing_from_latest_file(self):
        import_product_excel_rows(
            {
                "products": [self.product_row()],
                "options": [],
                "companies": [
                    self.company_row("Com-1", "Company One"),
                    self.company_row("Com-2", "Company Two"),
                ],
                "company_options": [
                    self.company_option_row("Com-1", "Size 42"),
                    self.company_option_row("Com-2", "Size 43"),
                ],
            }
        )

        import_product_excel_rows(
            {
                "products": [self.product_row()],
                "options": [],
                "companies": [self.company_row("Com-1", "Company One")],
                "company_options": [self.company_option_row("Com-1", "Size 44")],
            }
        )

        product = Product.objects.get(sku="SHOES-001")
        product_companies = ProductCompany.objects.filter(product=product).select_related("company")
        options = ProductCompanyOption.objects.filter(company__product=product)

        self.assertEqual([relation.company.code for relation in product_companies], ["Com-1"])
        self.assertEqual([option.name for option in options], ["Size 44"])
        self.assertFalse(ProductCompany.objects.filter(product=product, company__code="Com-2").exists())

    def test_import_removes_company_links_when_product_becomes_normal(self):
        self.import_rows("")

        import_product_excel_rows(
            {
                "products": [
                    self.product_row(
                        price=Decimal("20"),
                        product_type=Product.PRODUCT_TYPE_NORMAL,
                        has_options=False,
                    )
                ],
                "options": [],
                "companies": [],
                "company_options": [],
            }
        )

        product = Product.objects.get(sku="SHOES-001")

        self.assertEqual(product.product_type, Product.PRODUCT_TYPE_NORMAL)
        self.assertFalse(ProductCompany.objects.filter(product=product).exists())

    def test_import_removes_product_options_missing_from_latest_file(self):
        normal_product_row = self.product_row(
            price=Decimal("0"),
            product_type=Product.PRODUCT_TYPE_NORMAL,
            has_options=True,
        )
        import_product_excel_rows(
            {
                "products": [normal_product_row],
                "options": [self.option_row("Small"), self.option_row("Large")],
                "companies": [],
                "company_options": [],
            }
        )

        import_product_excel_rows(
            {
                "products": [normal_product_row],
                "options": [self.option_row("Small")],
                "companies": [],
                "company_options": [],
            }
        )

        product = Product.objects.get(sku="SHOES-001")

        self.assertEqual(
            list(ProductOption.objects.filter(product=product).values_list("name", flat=True)),
            ["Small"],
        )

    def test_company_is_shared_and_updated_by_stable_company_id(self):
        self.import_rows("https://example.com/old-logo.png")
        first_company = Company.objects.get(code="Com-1")
        product_company = ProductCompany.objects.select_related("company").get()
        second_product = Product.objects.create(
            category=self.category,
            name="Walking Shoes",
            product_type=Product.PRODUCT_TYPE_COMPANY_GROUPED,
            price=Decimal("0"),
            sku="SHOES-002",
        )
        second_product_company = ProductCompany.objects.create(
            product=second_product,
            company=first_company,
        )

        self.import_rows(
            "https://example.com/new-logo.png",
            company_name="Updated Company",
        )

        first_company.refresh_from_db()
        product_company.refresh_from_db()
        second_product_company.refresh_from_db()
        self.assertEqual(Company.objects.filter(code__iexact="Com-1").count(), 1)
        self.assertEqual(ProductCompany.objects.count(), 2)
        self.assertEqual(product_company.company_id, first_company.id)
        self.assertEqual(second_product_company.company_id, first_company.id)
        self.assertEqual(first_company.name, "Updated Company")
        self.assertEqual(
            first_company.external_logo_url,
            "https://example.com/new-logo.png",
        )
        self.assertEqual(product_company.company.external_logo_url, first_company.external_logo_url)
        self.assertEqual(
            second_product_company.company.external_logo_url,
            first_company.external_logo_url,
        )

    def test_empty_logo_url_uses_no_per_product_logo_copy(self):
        self.import_rows("")

        company = Company.objects.get(code="Com-1")
        product_company = ProductCompany.objects.select_related("company").get()
        self.assertFalse(company.logo)
        self.assertEqual(company.external_logo_url, "")
        self.assertFalse(hasattr(product_company, "logo_id"))

    def test_import_merges_matching_legacy_company_instead_of_duplicating_it(self):
        product = Product.objects.create(
            category=self.category,
            name="Running Shoes",
            product_type=Product.PRODUCT_TYPE_COMPANY_GROUPED,
            price=Decimal("0"),
            sku="SHOES-001",
        )
        legacy_company = Company.objects.create(
            code="legacy-1",
            name="Company One",
        )
        legacy_relation = ProductCompany.objects.create(
            product=product,
            company=legacy_company,
        )
        ProductCompanyOption.objects.create(
            company=legacy_relation,
            name="Size 42",
            price=Decimal("14.00"),
        )

        self.import_rows("https://example.com/new-logo.png")

        canonical_company = Company.objects.get(code="Com-1")
        relation = ProductCompany.objects.get(
            product=product,
            company=canonical_company,
        )
        self.assertFalse(Company.objects.filter(code__startswith="legacy-").exists())
        self.assertEqual(ProductCompany.objects.filter(product=product).count(), 1)
        self.assertEqual(relation.options.count(), 1)
        self.assertEqual(relation.options.get().price, Decimal("15.50"))

    def test_import_applies_company_price_link_override(self):
        self.import_rows("", is_price_linked_to_dollar=True)

        relation = ProductCompany.objects.get()

        self.assertIs(relation.is_price_linked_to_dollar, True)

    def test_excel_company_options_accept_price_link_override(self):
        workbook = Workbook()
        products_sheet = workbook.active
        products_sheet.title = "Products"
        products_sheet.append(PRODUCT_EXCEL_HEADERS)
        products_sheet.append(
            [
                "Running Shoes",
                "shoes",
                "",
                "",
                "",
                "company_grouped",
                "inherit",
                "",
                "inherit",
                "",
                "",
                "piece",
                "true",
                "false",
                "",
                "",
                "SHOES-001",
            ]
        )
        companies_sheet = workbook.create_sheet("Companies")
        companies_sheet.append(PRODUCT_COMPANY_EXCEL_HEADERS)
        companies_sheet.append(["Com-1", "Company One", "", "", "0", "true"])
        company_options_sheet = workbook.create_sheet("CompanyOptions")
        company_options_sheet.append(PRODUCT_COMPANY_OPTION_EXCEL_HEADERS)
        company_options_sheet.append(["SHOES-001", "Com-1", "true", "Size 42", "15.50", "true", "0"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        valid_rows, errors = validate_product_excel_workbook(output)

        self.assertEqual(errors, [])
        self.assertIs(valid_rows["company_options"][0]["is_price_linked_to_dollar"], True)


class DashboardOrderRejectionTests(TestCase):
    def setUp(self):
        self.staff_user = get_user_model().objects.create_user(
            username="order-manager",
            password="test-password",
            is_staff=True,
        )
        customer_user = get_user_model().objects.create_user(
            username="rejected-customer",
            password="test-password",
        )
        self.profile = CustomerProfile.objects.create(
            user=customer_user,
            full_name="Rejected Customer",
            phone_number="944444443",
        )
        self.client.force_login(self.staff_user)
        session = self.client.session
        session["site_language"] = "en"
        session.save()

    def create_order(self, status=CustomerOrder.STATUS_BEING_PREPARED, invoice_number="AR-REJECT-1"):
        return CustomerOrder.objects.create(
            profile=self.profile,
            invoice_number=invoice_number,
            service_type=CustomerOrder.SERVICE_PICKUP,
            status=status,
            address_snapshot="Pickup",
        )

    def test_staff_can_reject_active_order(self):
        order = self.create_order()

        response = self.client.post(
            reverse("dashboard:pending-order-reject", kwargs={"pk": order.pk})
        )

        self.assertRedirects(response, reverse("dashboard:pending-orders"))
        order.refresh_from_db()
        self.assertEqual(order.status, CustomerOrder.STATUS_CANCELLED)
        self.assertIsNotNone(order.completed_at)

    def test_staff_cannot_reject_completed_order(self):
        order = self.create_order(status=CustomerOrder.STATUS_DONE)

        response = self.client.post(
            reverse("dashboard:pending-order-reject", kwargs={"pk": order.pk})
        )

        self.assertRedirects(
            response,
            reverse("dashboard:pending-order-detail", kwargs={"pk": order.pk}),
        )
        order.refresh_from_db()
        self.assertEqual(order.status, CustomerOrder.STATUS_DONE)

    def test_non_staff_cannot_reject_order(self):
        non_staff = get_user_model().objects.create_user(
            username="not-manager",
            password="test-password",
        )
        self.client.force_login(non_staff)
        order = self.create_order()

        response = self.client.post(
            reverse("dashboard:pending-order-reject", kwargs={"pk": order.pk})
        )

        self.assertEqual(response.status_code, 403)
        order.refresh_from_db()
        self.assertEqual(order.status, CustomerOrder.STATUS_BEING_PREPARED)

    def test_pending_orders_render_reject_action(self):
        order = self.create_order()

        response = self.client.get(reverse("dashboard:pending-orders"))

        self.assertContains(
            response,
            reverse("dashboard:pending-order-reject", kwargs={"pk": order.pk}),
        )
        self.assertContains(response, "Reject order")


class ManagerNewOrderAlertTests(TestCase):
    def setUp(self):
        self.staff_user = get_user_model().objects.create_user(
            username="alert-manager",
            password="test-password",
            is_staff=True,
        )
        customer_user = get_user_model().objects.create_user(
            username="alert-customer",
            password="test-password",
        )
        self.profile = CustomerProfile.objects.create(
            user=customer_user,
            full_name="Alert Customer",
            phone_number="944444445",
        )
        self.client.force_login(self.staff_user)
        session = self.client.session
        session["site_language"] = "en"
        session.save()

    def create_order(self, **kwargs):
        defaults = {
            "profile": self.profile,
            "invoice_number": "AR-ALERT-1",
            "service_type": CustomerOrder.SERVICE_PICKUP,
            "status": CustomerOrder.STATUS_BEING_PREPARED,
            "address_snapshot": "Pickup",
            "total_min": Decimal("1000"),
            "total_max": Decimal("1000"),
        }
        defaults.update(kwargs)
        return CustomerOrder.objects.create(**defaults)

    def test_check_new_returns_unseen_active_orders(self):
        order = self.create_order()
        self.create_order(invoice_number="AR-ALERT-SEEN", manager_seen=True)
        self.create_order(invoice_number="AR-ALERT-DONE", status=CustomerOrder.STATUS_DONE)

        response = self.client.get(reverse("dashboard:orders-check-new"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["orders"]), 1)
        self.assertEqual(payload["orders"][0]["order_id"], order.pk)
        self.assertEqual(payload["orders"][0]["order_number"], "AR-ALERT-1")
        self.assertEqual(payload["orders"][0]["customer_name"], "Alert Customer")
        self.assertEqual(
            payload["orders"][0]["detail_url"],
            reverse("dashboard:pending-order-detail", kwargs={"pk": order.pk}),
        )

    def test_acknowledge_marks_order_seen(self):
        order = self.create_order()

        response = self.client.post(reverse("dashboard:orders-check-new"), {"order_id": order.pk})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        order.refresh_from_db()
        self.assertTrue(order.manager_seen)

    def test_non_staff_cannot_check_new_orders(self):
        non_staff = get_user_model().objects.create_user(
            username="not-alert-manager",
            password="test-password",
        )
        self.client.force_login(non_staff)

        response = self.client.get(reverse("dashboard:orders-check-new"))

        self.assertEqual(response.status_code, 403)


class ExcelImageImportTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name="Footwear", slug="shoes")

    @staticmethod
    def png_bytes():
        output = BytesIO()
        Image.new("RGB", (2, 2), color="red").save(output, format="PNG")
        return output.getvalue()

    @staticmethod
    def append_dict_row(worksheet, headers, values):
        worksheet.append([values.get(header, "") for header in headers])

    def build_workbook(
        self,
        image_value,
        external_url="",
        company_logo=None,
        category_image=None,
    ):
        company_logo = image_value if company_logo is None else company_logo
        category_image = image_value if category_image is None else category_image
        workbook = Workbook()
        products_sheet = workbook.active
        products_sheet.title = "Products"
        products_sheet.append(PRODUCT_EXCEL_HEADERS)
        self.append_dict_row(
            products_sheet,
            PRODUCT_EXCEL_HEADERS,
            {
                "product_name": "Running Shoes",
                "category": "shoes",
                "price": "15.50",
                "product_type": "company_grouped",
                "image": image_value,
                "external_image_url": external_url,
                "sku": "SHOES-001",
                "is_available": "true",
            },
        )

        companies_sheet = workbook.create_sheet("Companies")
        companies_sheet.append(PRODUCT_COMPANY_EXCEL_HEADERS)
        self.append_dict_row(
            companies_sheet,
            PRODUCT_COMPANY_EXCEL_HEADERS,
            {
                "company_id": "Com-1",
                "company_name": "Company One",
                "logo": company_logo,
                "external_logo_url": external_url,
                "is_active": "true",
            },
        )

        company_options_sheet = workbook.create_sheet("CompanyOptions")
        company_options_sheet.append(PRODUCT_COMPANY_OPTION_EXCEL_HEADERS)
        self.append_dict_row(
            company_options_sheet,
            PRODUCT_COMPANY_OPTION_EXCEL_HEADERS,
            {
                "product_sku": "SHOES-001",
                "company_id": "Com-1",
                "option_name": "Size 42",
                "price": "15.50",
                "is_available": "true",
            },
        )

        categories_sheet = workbook.create_sheet("Categories")
        categories_sheet.append(CATEGORY_IMAGE_EXCEL_HEADERS)
        self.append_dict_row(
            categories_sheet,
            CATEGORY_IMAGE_EXCEL_HEADERS,
            {
                "category": "shoes",
                "name": "Footwear",
                "slug": "shoes",
                "image": category_image,
                "external_image_url": external_url,
            },
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def test_server_import_folders_populate_main_image_fields(self):
        with TemporaryDirectory() as temporary_directory:
            media_root = Path(temporary_directory) / "media"
            product_image = media_root / "import" / "products" / "apple.png"
            company_logo = media_root / "import" / "companies" / "adidas.png"
            category_image = media_root / "import" / "categories" / "vegetables.png"
            for image_path in (product_image, company_logo, category_image):
                image_path.parent.mkdir(parents=True, exist_ok=True)
                image_path.write_bytes(self.png_bytes())
            workbook = self.build_workbook(
                "apple.png",
                company_logo="adidas.png",
                category_image="vegetables.png",
            )

            with override_settings(MEDIA_ROOT=media_root):
                valid_rows, errors = validate_product_excel_workbook(workbook)
                self.assertEqual(errors, [])
                import_product_excel_rows(valid_rows)

                product = Product.objects.get(sku="SHOES-001")
                company = Company.objects.get(code="Com-1")
                self.category.refresh_from_db()

                self.assertTrue(product.primary_image.name.startswith("products/"))
                self.assertEqual(product.external_image_url, "")
                self.assertTrue(company.logo.name.startswith("product-companies/"))
                self.assertEqual(company.external_logo_url, "")
                self.assertTrue(self.category.cover_image.name.startswith("categories/"))
                self.assertEqual(self.category.external_image_url, "")
                self.assertTrue((media_root / product.primary_image.name).is_file())
                self.assertTrue((media_root / company.logo.name).is_file())
                self.assertTrue((media_root / self.category.cover_image.name).is_file())

    def test_missing_server_images_log_warnings_and_do_not_stop_import(self):
        with TemporaryDirectory() as temporary_directory:
            media_root = Path(temporary_directory) / "media"
            workbook = self.build_workbook(
                "missing-product.png",
                company_logo="missing-company.png",
                category_image="missing-category.png",
            )

            with override_settings(MEDIA_ROOT=media_root):
                with self.assertLogs("apps.dashboard.views", level="WARNING") as captured:
                    valid_rows, errors = validate_product_excel_workbook(workbook)
                self.assertEqual(errors, [])
                self.assertEqual(len(captured.output), 3)
                import_product_excel_rows(valid_rows)

                product = Product.objects.get(sku="SHOES-001")
                company = Company.objects.get(code="Com-1")
                self.category.refresh_from_db()

                self.assertFalse(product.primary_image)
                self.assertFalse(company.logo)
                self.assertFalse(self.category.cover_image)

    def test_explicit_external_columns_remain_external_fallbacks(self):
        external_url = "https://example.com/external.png"
        workbook = self.build_workbook("", external_url=external_url)

        valid_rows, errors = validate_product_excel_workbook(workbook)
        self.assertEqual(errors, [])
        import_product_excel_rows(valid_rows)

        product = Product.objects.get(sku="SHOES-001")
        company = Company.objects.get(code="Com-1")
        self.category.refresh_from_db()

        self.assertFalse(product.primary_image)
        self.assertEqual(product.external_image_url, external_url)
        self.assertFalse(company.logo)
        self.assertEqual(company.external_logo_url, external_url)
        self.assertFalse(self.category.cover_image)
        self.assertEqual(self.category.external_image_url, external_url)

    def test_url_image_is_downloaded_and_validated(self):
        content = self.png_bytes()

        class FakeResponse(BytesIO):
            headers = {"Content-Length": str(len(content))}

            def geturl(self):
                return "https://example.com/images/logo.png"

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc_value, traceback):
                self.close()

        with patch("apps.dashboard.views.urlopen", return_value=FakeResponse(content)):
            image_data = load_excel_image(
                "https://example.com/logo.png",
                [],
                import_directory="companies",
            )

        self.assertEqual(image_data["name"], "logo.png")
        self.assertEqual(image_data["content"], content)
